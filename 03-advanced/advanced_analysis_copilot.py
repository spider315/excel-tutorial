"""
03-advanced/advanced_analysis.py
多維度銷售分析報表產生器

任務一：資料清理與標準化
任務二：多維度銷售分析 → sales_analysis_report.xlsx
任務三：預算達成率 → kpi_dashboard.xlsx
任務四：產品利潤交叉分析 → product_profit_report.xlsx
任務五：資料品質報告 → data_quality_report.xlsx
"""

import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, FormulaRule,
    CellIsRule, Rule
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# ── 路徑設定 ──────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR  = os.path.join(BASE_DIR, "raw")
OUT_DIR  = os.path.join(BASE_DIR, "output")
os.makedirs(OUT_DIR, exist_ok=True)

# ── 樣式常數 ──────────────────────────────────────────────────
BLUE_HEADER   = "2F5496"
GREEN_HEADER  = "548235"
ORANGE_HEADER = "BF8F00"

FILL_BLUE   = PatternFill("solid", fgColor=BLUE_HEADER)
FILL_GREEN  = PatternFill("solid", fgColor=GREEN_HEADER)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE_HEADER)

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD   = Font(bold=True, size=11)

THIN_SIDE   = Side(style="thin", color="BBBBBB")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE,  bottom=THIN_SIDE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

# KPI conditional format colors
FILL_KPI_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_KPI_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_KPI_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_TOP3       = PatternFill("solid", fgColor="C6EFCE")


# ══════════════════════════════════════════════════════════════
#  工具函式
# ══════════════════════════════════════════════════════════════

def col_width(text: str) -> float:
    """中文字算 2 字寬，英數字算 1 字寬。"""
    w = 0.0
    for ch in str(text):
        w += 2.0 if '\u4e00' <= ch <= '\u9fff' else 1.0
    return w


def auto_col_widths(ws, min_w=10, max_w=30):
    """自動調整工作表欄寬（中文字算 2 字寬）。"""
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, col_width(str(cell.value)))
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


def style_header_row(ws, row_num: int, fill: PatternFill = None):
    """設定指定列為標題列樣式（粗體白字、深色底、置中換行）。"""
    if fill is None:
        fill = FILL_BLUE
    for cell in ws[row_num]:
        cell.font      = FONT_HEADER
        cell.fill      = fill
        cell.alignment = ALIGN_CENTER
        cell.border    = THIN_BORDER


def style_data_rows(ws, start_row: int, end_row: int = None):
    """套用細邊框 + 垂直置中到資料列。"""
    if end_row is None:
        end_row = ws.max_row
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def write_df_to_sheet(ws, df: pd.DataFrame, header_fill: PatternFill = None,
                      start_row: int = 1):
    """將 DataFrame 寫入工作表，自動套用標題與資料樣式。"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    style_header_row(ws, start_row, fill=header_fill)
    style_data_rows(ws, start_row + 1, start_row + len(df))
    auto_col_widths(ws)


# ══════════════════════════════════════════════════════════════
#  任務一：資料清理
# ══════════════════════════════════════════════════════════════

def clean_data():
    """讀取四份原始 Excel，清理後輸出，並記錄 cleaning_log。"""
    log_rows = []

    def log(source, row_num, col, orig, fixed, desc):
        log_rows.append({
            "來源檔案": source, "列號": row_num, "欄位": col,
            "原始值": orig, "修正值": fixed, "問題描述": desc
        })

    # ── 1-A. monthly_sales ──────────────────────────────────
    fname = "monthly_sales.xlsx"
    df = pd.read_excel(os.path.join(RAW_DIR, fname))

    for i, row in df.iterrows():
        row_num = i + 2  # Excel 列號（含標題列）

        # 日期格式：/ → -
        date_str = str(row["訂單日期"])
        if "/" in date_str:
            fixed = date_str.replace("/", "-")
            log(fname, row_num, "訂單日期", date_str, fixed, "日期格式含斜線")
            df.at[i, "訂單日期"] = fixed

        # 業務員：去空白、移除括號標注
        orig_name = str(row["業務員"])
        fixed_name = re.sub(r'[（(][^）)]*[）)]', '', orig_name).strip()
        fixed_name = re.sub(r'\s+', '', fixed_name)
        if fixed_name != orig_name:
            log(fname, row_num, "業務員", orig_name, fixed_name, "業務員名稱含空白或括號")
            df.at[i, "業務員"] = fixed_name

        # 產品名稱：去空白、修正錯字
        orig_prod = str(row["產品名稱"])
        fixed_prod = orig_prod.strip()
        fixed_prod = re.sub(r'\s+', ' ', fixed_prod)        # 多餘內部空白
        fixed_prod = fixed_prod.replace("智彗", "智慧")
        fixed_prod = fixed_prod.replace("藍芽", "藍牙")
        if fixed_prod != orig_prod:
            desc_parts = []
            if orig_prod.strip() != orig_prod or re.search(r'\s{2,}', orig_prod):
                desc_parts.append("產品名稱含多餘空白")
            if "智彗" in orig_prod:
                desc_parts.append("錯字智彗→智慧")
            if "藍芽" in orig_prod:
                desc_parts.append("錯字藍芽→藍牙")
            if re.search(r'(?<!\s)\s(?!\s)', orig_prod) and orig_prod.strip() == orig_prod:
                desc_parts.append("產品名稱含不當內部空格")
            if not desc_parts:
                desc_parts.append("產品名稱格式修正")
            log(fname, row_num, "產品名稱", orig_prod, fixed_prod, "；".join(desc_parts))
            df.at[i, "產品名稱"] = fixed_prod

    # 數值型態
    for col in ["數量", "單價", "銷售金額"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df.to_excel(os.path.join(OUT_DIR, "cleaned_monthly_sales.xlsx"), index=False)
    print(f"✔ cleaned_monthly_sales.xlsx 完成（{len(df)} 筆）")

    # ── 1-B. budget_targets ─────────────────────────────────
    fname = "budget_targets.xlsx"
    df_budget = pd.read_excel(os.path.join(RAW_DIR, fname))

    for i, row in df_budget.iterrows():
        row_num = i + 2
        # 業務員名稱去空白 / 括號
        orig_name = str(row["業務員"])
        fixed_name = re.sub(r'[（(][^）)]*[）)]', '', orig_name).strip()
        fixed_name = re.sub(r'\s+', '', fixed_name)
        if fixed_name != orig_name:
            log(fname, row_num, "業務員", orig_name, fixed_name, "業務員名稱含空白或括號")
            df_budget.at[i, "業務員"] = fixed_name

    for col in ["年度目標金額", "Q1目標", "Q2目標", "Q3目標", "Q4目標"]:
        df_budget[col] = pd.to_numeric(df_budget[col], errors="coerce")

    df_budget.to_excel(os.path.join(OUT_DIR, "cleaned_budget_targets.xlsx"), index=False)
    print(f"✔ cleaned_budget_targets.xlsx 完成（{len(df_budget)} 筆）")

    # ── 1-C. customer_feedback ──────────────────────────────
    fname = "customer_feedback.xlsx"
    df_fb = pd.read_excel(os.path.join(RAW_DIR, fname))

    for i, row in df_fb.iterrows():
        row_num = i + 2
        # 日期格式
        date_str = str(row["日期"])
        if "/" in date_str:
            fixed = date_str.replace("/", "-")
            log(fname, row_num, "日期", date_str, fixed, "日期格式含斜線")
            df_fb.at[i, "日期"] = fixed
        # 產品名稱
        orig_prod = str(row["產品名稱"])
        fixed_prod = orig_prod.strip()
        fixed_prod = re.sub(r'\s+', ' ', fixed_prod)
        fixed_prod = fixed_prod.replace("智彗", "智慧")
        fixed_prod = fixed_prod.replace("藍芽", "藍牙")
        if fixed_prod != orig_prod:
            log(fname, row_num, "產品名稱", orig_prod, fixed_prod, "產品名稱格式修正")
            df_fb.at[i, "產品名稱"] = fixed_prod

    df_fb["滿意度評分"] = pd.to_numeric(df_fb["滿意度評分"], errors="coerce")
    df_fb.to_excel(os.path.join(OUT_DIR, "cleaned_customer_feedback.xlsx"), index=False)
    print(f"✔ cleaned_customer_feedback.xlsx 完成（{len(df_fb)} 筆）")

    # ── 1-D. cleaning_log ───────────────────────────────────
    df_log = pd.DataFrame(log_rows, columns=["來源檔案", "列號", "欄位", "原始值", "修正值", "問題描述"])

    wb_log = Workbook()
    ws_log = wb_log.active
    ws_log.title = "清理日誌"
    write_df_to_sheet(ws_log, df_log)
    wb_log.save(os.path.join(OUT_DIR, "cleaning_log.xlsx"))
    print(f"✔ cleaning_log.xlsx 完成（{len(df_log)} 筆修正紀錄）")

    return df, df_budget, df_fb, df_log


# ══════════════════════════════════════════════════════════════
#  任務二：多維度銷售分析
# ══════════════════════════════════════════════════════════════

def build_sales_analysis(df_sales: pd.DataFrame):
    """產生 sales_analysis_report.xlsx（5 個 Sheet）。"""

    # 前置：排除銷售金額 ≤ 0 或數量 ≤ 0
    df = df_sales[
        (pd.to_numeric(df_sales["銷售金額"], errors="coerce") > 0) &
        (pd.to_numeric(df_sales["數量"], errors="coerce") > 0)
    ].copy()
    df["訂單日期"] = pd.to_datetime(df["訂單日期"], errors="coerce")
    df["月份"] = df["訂單日期"].dt.to_period("M")
    df["銷售金額"] = pd.to_numeric(df["銷售金額"], errors="coerce")
    df["數量"]     = pd.to_numeric(df["數量"],    errors="coerce")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: 月度銷售趨勢 ─────────────────────────────
    ws1 = wb.create_sheet("月度銷售趨勢")

    monthly = (
        df.groupby("月份")
          .agg(訂單數=("訂單編號", "count"),
               銷售總額=("銷售金額", "sum"))
          .reset_index()
    )
    monthly["月份"] = monthly["月份"].astype(str)
    monthly["平均客單價"] = (monthly["銷售總額"] / monthly["訂單數"]).round(0)
    monthly["月成長率%"] = (
        monthly["銷售總額"].pct_change() * 100
    ).round(2)

    headers = ["月份", "訂單數", "銷售總額", "平均客單價", "月成長率%"]
    ws1.append(headers)
    style_header_row(ws1, 1)

    for _, row in monthly.iterrows():
        ws1.append([
            row["月份"], int(row["訂單數"]),
            round(float(row["銷售總額"]), 0),
            round(float(row["平均客單價"]), 0),
            row["月成長率%"] if pd.notna(row["月成長率%"]) else ""
        ])
    style_data_rows(ws1, 2, 1 + len(monthly))
    auto_col_widths(ws1)

    n_rows = len(monthly)

    # 折線圖：銷售總額（放在 A15）
    line_chart = LineChart()
    line_chart.title = "月度銷售總額趨勢"
    line_chart.style = 10
    line_chart.y_axis.title = "銷售金額"
    line_chart.x_axis.title = "月份"
    line_chart.width  = 20
    line_chart.height = 12
    data_ref  = Reference(ws1, min_col=3, min_row=1, max_row=1 + n_rows)
    cats_ref  = Reference(ws1, min_col=1, min_row=2, max_row=1 + n_rows)
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    ws1.add_chart(line_chart, "A15")

    # 長條圖：訂單數（放在 A32）
    bar_chart = BarChart()
    bar_chart.title = "月度訂單數"
    bar_chart.style = 10
    bar_chart.y_axis.title = "訂單數"
    bar_chart.x_axis.title = "月份"
    bar_chart.width  = 20
    bar_chart.height = 12
    data_ref2 = Reference(ws1, min_col=2, min_row=1, max_row=1 + n_rows)
    bar_chart.add_data(data_ref2, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    ws1.add_chart(bar_chart, "A32")

    # ── Sheet 2: 部門銷售樞紐 ─────────────────────────────
    ws2 = wb.create_sheet("部門銷售樞紐")

    pivot = df.pivot_table(
        index="部門", columns="月份", values="銷售金額",
        aggfunc="sum", fill_value=0
    )
    pivot.columns = [str(c) for c in pivot.columns]
    pivot["年度合計"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    for r_idx, row in enumerate(dataframe_to_rows(pivot, index=False, header=True), 1):
        ws2.append(row)
    style_header_row(ws2, 1, fill=FILL_GREEN)
    style_data_rows(ws2, 2, len(pivot) + 1)
    auto_col_widths(ws2)

    # ── Sheet 3: 業務員排名 ───────────────────────────────
    ws3 = wb.create_sheet("業務員排名")

    rank_df = (
        df.groupby(["業務員", "部門"])
          .agg(訂單數=("訂單編號", "count"),
               銷售總額=("銷售金額", "sum"))
          .reset_index()
    )
    rank_df["平均每單金額"] = (rank_df["銷售總額"] / rank_df["訂單數"]).round(0)
    rank_df = rank_df.sort_values("銷售總額", ascending=False).reset_index(drop=True)
    rank_df.insert(0, "排名", range(1, len(rank_df) + 1))

    cols3 = ["排名", "業務員", "部門", "訂單數", "銷售總額", "平均每單金額"]
    rank_df = rank_df[cols3]

    for r_idx, row in enumerate(dataframe_to_rows(rank_df, index=False, header=True), 1):
        ws3.append(row)
    style_header_row(ws3, 1)
    style_data_rows(ws3, 2, len(rank_df) + 1)

    # Top 3 淺綠底
    for row_num in range(2, 5):
        for cell in ws3[row_num]:
            cell.fill = FILL_TOP3

    auto_col_widths(ws3)

    # ── Sheet 4: 產品類別分布 ─────────────────────────────
    ws4 = wb.create_sheet("產品類別分布")

    # 從 product_catalog 取得類別對應
    cat = pd.read_excel(os.path.join(RAW_DIR, "product_catalog.xlsx"))
    cat["產品名稱_clean"] = cat["產品名稱"].str.strip().str.replace(r'\s+', ' ', regex=True)
    df["產品名稱_clean"]  = df["產品名稱"].str.strip().str.replace(r'\s+', ' ', regex=True)
    df_merged = df.merge(
        cat[["產品編號", "產品類別"]],
        on="產品編號", how="left"
    )

    cat_df = (
        df_merged.groupby("產品類別")
                 .agg(銷售筆數=("訂單編號", "count"),
                      銷售總額=("銷售金額", "sum"))
                 .reset_index()
    )
    total4 = cat_df["銷售總額"].sum()
    cat_df["銷售佔比"] = (cat_df["銷售總額"] / total4 * 100).round(2).astype(str) + "%"
    cat_df = cat_df.sort_values("銷售總額", ascending=False).reset_index(drop=True)

    for r_idx, row in enumerate(dataframe_to_rows(cat_df, index=False, header=True), 1):
        ws4.append(row)
    style_header_row(ws4, 1, fill=FILL_ORANGE)
    style_data_rows(ws4, 2, len(cat_df) + 1)
    auto_col_widths(ws4)

    # 圓餅圖（A10）
    pie = PieChart()
    pie.title = "產品類別銷售佔比"
    pie.style = 10
    pie.width  = 18
    pie.height = 14
    n4 = len(cat_df)
    labels_ref = Reference(ws4, min_col=1, min_row=2, max_row=1 + n4)
    data_ref4  = Reference(ws4, min_col=3, min_row=1, max_row=1 + n4)
    pie.add_data(data_ref4, titles_from_data=True)
    pie.dataLabels = None
    pie.set_categories(labels_ref)
    ws4.add_chart(pie, "A10")

    # ── Sheet 5: 客戶區域分布 ─────────────────────────────
    ws5 = wb.create_sheet("客戶區域分布")

    region_df = (
        df.groupby("客戶區域")
          .agg(銷售筆數=("訂單編號", "count"),
               銷售總額=("銷售金額", "sum"))
          .reset_index()
          .sort_values("銷售總額", ascending=False)
          .reset_index(drop=True)
    )

    for r_idx, row in enumerate(dataframe_to_rows(region_df, index=False, header=True), 1):
        ws5.append(row)
    style_header_row(ws5, 1, fill=FILL_BLUE)
    style_data_rows(ws5, 2, len(region_df) + 1)
    auto_col_widths(ws5)

    wb.save(os.path.join(OUT_DIR, "sales_analysis_report.xlsx"))
    print("✔ sales_analysis_report.xlsx 完成（5 sheets）")

    return df, df_merged


# ══════════════════════════════════════════════════════════════
#  任務三：預算達成率 KPI
# ══════════════════════════════════════════════════════════════

def build_kpi_dashboard(df_sales_clean: pd.DataFrame, df_budget: pd.DataFrame):
    """產生 kpi_dashboard.xlsx。"""

    df = df_sales_clean[
        (df_sales_clean["銷售金額"] > 0) &
        (df_sales_clean["數量"]     > 0)
    ].copy()

    actual = (
        df.groupby("業務員")
          .agg(實際銷售額=("銷售金額", "sum"),
               部門=("部門", "first"))
          .reset_index()
    )

    # 合併預算
    kpi = actual.merge(
        df_budget[["業務員", "年度目標金額"]],
        on="業務員", how="outer"
    ).fillna(0)
    kpi = kpi.rename(columns={"年度目標金額": "目標金額"})
    kpi["實際銷售額"] = pd.to_numeric(kpi["實際銷售額"], errors="coerce").fillna(0)
    kpi["目標金額"]   = pd.to_numeric(kpi["目標金額"],   errors="coerce").fillna(0)
    kpi["達成率%"] = np.where(
        kpi["目標金額"] > 0,
        (kpi["實際銷售額"] / kpi["目標金額"] * 100).round(2),
        0.0
    )

    def kpi_status(pct):
        if pct >= 120: return "★超標"
        if pct >= 100: return "✔達成"
        if pct >= 80:  return "△接近"
        return "✘未達成"

    kpi["狀態"] = kpi["達成率%"].apply(kpi_status)
    kpi = kpi.sort_values("達成率%", ascending=False).reset_index(drop=True)
    kpi_out = kpi[["業務員", "部門", "實際銷售額", "目標金額", "達成率%", "狀態"]]

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: 個人 KPI ───────────────────────────────
    ws1 = wb.create_sheet("個人KPI達成率")

    for r_idx, row in enumerate(dataframe_to_rows(kpi_out, index=False, header=True), 1):
        ws1.append(row)
    style_header_row(ws1, 1)
    style_data_rows(ws1, 2, len(kpi_out) + 1)
    auto_col_widths(ws1)

    # 條件格式（依達成率% 欄 = 第 5 欄 E）
    n_kpi = len(kpi_out)
    pct_col = "E"
    cell_range = f"{pct_col}2:{pct_col}{1 + n_kpi}"

    # ≥100 綠底
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")

    ws1.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["100"], fill=green_fill)
    )
    # 80-99 黃底
    ws1.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["80", "99.99"], fill=yellow_fill)
    )
    # <80 紅底
    ws1.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["80"], fill=red_fill)
    )

    # ── Sheet 2: 部門 KPI 彙總 ──────────────────────────
    ws2 = wb.create_sheet("部門KPI彙總")

    dept_kpi = (
        kpi.groupby("部門")
           .agg(人數=("業務員", "count"),
                實際銷售額=("實際銷售額", "sum"),
                目標金額=("目標金額", "sum"))
           .reset_index()
    )
    dept_kpi["達成率%"] = np.where(
        dept_kpi["目標金額"] > 0,
        (dept_kpi["實際銷售額"] / dept_kpi["目標金額"] * 100).round(2),
        0.0
    )
    dept_kpi = dept_kpi.sort_values("達成率%", ascending=False).reset_index(drop=True)

    for r_idx, row in enumerate(dataframe_to_rows(dept_kpi, index=False, header=True), 1):
        ws2.append(row)
    style_header_row(ws2, 1, fill=FILL_GREEN)
    style_data_rows(ws2, 2, len(dept_kpi) + 1)
    auto_col_widths(ws2)

    # 長條圖（A10）
    bar2 = BarChart()
    bar2.type   = "col"
    bar2.title  = "部門KPI達成率%"
    bar2.style  = 10
    bar2.y_axis.title = "達成率%"
    bar2.width  = 18
    bar2.height = 12
    n_dept = len(dept_kpi)
    data_ref = Reference(ws2, min_col=5, min_row=1, max_row=1 + n_dept)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=1 + n_dept)
    bar2.add_data(data_ref, titles_from_data=True)
    bar2.set_categories(cats_ref)
    ws2.add_chart(bar2, "A10")

    wb.save(os.path.join(OUT_DIR, "kpi_dashboard.xlsx"))
    print("✔ kpi_dashboard.xlsx 完成（2 sheets）")

    return kpi


# ══════════════════════════════════════════════════════════════
#  任務四：產品利潤交叉分析
# ══════════════════════════════════════════════════════════════

def build_product_profit(df_merged: pd.DataFrame):
    """產生 product_profit_report.xlsx。"""

    df = df_merged[
        (df_merged["銷售金額"] > 0) &
        (df_merged["數量"]     > 0)
    ].copy()

    # 從 product_catalog 取得成本
    cat = pd.read_excel(os.path.join(RAW_DIR, "product_catalog.xlsx"))

    # 從 customer_feedback 計算平均滿意度
    fb = pd.read_excel(os.path.join(OUT_DIR, "cleaned_customer_feedback.xlsx"))
    fb["滿意度評分"] = pd.to_numeric(fb["滿意度評分"], errors="coerce")
    fb_valid = fb[fb["滿意度評分"].between(1, 5)]
    avg_score = (
        fb_valid.groupby("產品編號")["滿意度評分"]
                .mean()
                .round(2)
                .reset_index()
                .rename(columns={"滿意度評分": "平均滿意度"})
    )

    # 銷售彙總
    sales_sum = (
        df.groupby(["產品編號", "產品名稱", "產品類別"])
          .agg(銷售筆數=("訂單編號", "count"),
               銷售數量=("數量", "sum"),
               銷售總額=("銷售金額", "sum"))
          .reset_index()
    )

    # merge 成本
    profit = sales_sum.merge(
        cat[["產品編號", "成本"]],
        on="產品編號", how="left"
    )
    profit["成本"] = pd.to_numeric(profit["成本"], errors="coerce").fillna(0)

    # merge 滿意度
    profit = profit.merge(avg_score, on="產品編號", how="left")

    # 毛利計算
    profit["總成本"] = (profit["銷售數量"] * profit["成本"]).round(0)
    profit["毛利"]   = (profit["銷售總額"] - profit["總成本"]).round(0)
    profit["毛利率%"] = np.where(
        profit["銷售總額"] > 0,
        ((profit["銷售總額"] - profit["總成本"]) / profit["銷售總額"] * 100).round(2),
        0.0
    )

    profit = profit.sort_values("毛利", ascending=False).reset_index(drop=True)

    out_cols = ["產品編號", "產品名稱", "產品類別",
                "銷售筆數", "銷售數量", "銷售總額",
                "成本", "總成本", "毛利", "毛利率%", "平均滿意度"]
    profit_out = profit[out_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "產品利潤分析"

    for r_idx, row in enumerate(dataframe_to_rows(profit_out, index=False, header=True), 1):
        ws.append(row)
    style_header_row(ws, 1, fill=FILL_GREEN)
    style_data_rows(ws, 2, len(profit_out) + 1)
    auto_col_widths(ws)

    # 毛利率% 欄加資料條（藍 #5B9BD5）
    # 找毛利率% 欄的欄號
    margin_col_idx = out_cols.index("毛利率%") + 1
    margin_col_letter = get_column_letter(margin_col_idx)
    margin_range = f"{margin_col_letter}2:{margin_col_letter}{1 + len(profit_out)}"

    ws.conditional_formatting.add(
        margin_range,
        DataBarRule(
            start_type="min", start_value=None,
            end_type="max",   end_value=None,
            color="5B9BD5"
        )
    )

    wb.save(os.path.join(OUT_DIR, "product_profit_report.xlsx"))
    print("✔ product_profit_report.xlsx 完成（1 sheet）")

    return profit


# ══════════════════════════════════════════════════════════════
#  任務五：資料品質報告
# ══════════════════════════════════════════════════════════════

def build_data_quality(df_sales_raw: pd.DataFrame,
                       df_budget_raw: pd.DataFrame,
                       df_fb_raw: pd.DataFrame,
                       df_log: pd.DataFrame):
    """產生 data_quality_report.xlsx（3 sheets）。"""

    issues = []

    def add_issue(itype, source, row_num, col, val, desc):
        issues.append({
            "問題類型": itype, "來源檔案": source,
            "列號": row_num, "欄位": col,
            "問題值": str(val), "問題描述": desc
        })

    # ── 負數金額 ───────────────────────────────────────────
    fname_s = "monthly_sales.xlsx"
    for i, row in df_sales_raw.iterrows():
        amt = pd.to_numeric(row.get("銷售金額"), errors="coerce")
        if pd.notna(amt) and amt < 0:
            add_issue("負數金額", fname_s, i + 2, "銷售金額", amt, "銷售金額為負數")

    # ── 數量 = 0 ───────────────────────────────────────────
    for i, row in df_sales_raw.iterrows():
        qty = pd.to_numeric(row.get("數量"), errors="coerce")
        if pd.notna(qty) and qty == 0:
            add_issue("數量為零", fname_s, i + 2, "數量", qty, "訂單數量為 0")

    # ── 重複訂單 ────────────────────────────────────────────
    dup_mask = df_sales_raw.duplicated(subset=["訂單編號"], keep=False)
    for i, row in df_sales_raw[dup_mask].iterrows():
        add_issue("重複訂單", fname_s, i + 2, "訂單編號",
                  row["訂單編號"], "訂單編號重複出現")

    # ── 評分超範圍 ──────────────────────────────────────────
    fname_fb = "customer_feedback.xlsx"
    for i, row in df_fb_raw.iterrows():
        score = pd.to_numeric(row.get("滿意度評分"), errors="coerce")
        if pd.notna(score) and not (1 <= score <= 5):
            add_issue("評分超範圍", fname_fb, i + 2, "滿意度評分",
                      score, f"滿意度評分 {score} 超出 1-5 範圍")

    # ── 預算 Q 合計超 115% ──────────────────────────────────
    fname_b = "budget_targets.xlsx"
    for i, row in df_budget_raw.iterrows():
        annual = pd.to_numeric(row.get("年度目標金額"), errors="coerce")
        q_sum  = sum(pd.to_numeric(row.get(f"Q{q}目標", 0), errors="coerce")
                     for q in range(1, 5))
        if pd.notna(annual) and annual > 0 and (q_sum / annual) > 1.15:
            ratio = round(q_sum / annual * 100, 2)
            add_issue("預算Q合計超115%", fname_b, i + 2, "Q1~Q4目標",
                      f"Q合計={q_sum}", f"Q1-Q4合計佔年度目標 {ratio}%，超過 115%")

    # ── Sheet 1: 問題清單 ──────────────────────────────────
    df_issues = pd.DataFrame(issues)

    # ── Sheet 2: 問題統計 ──────────────────────────────────
    if not df_issues.empty:
        df_summary = df_issues.groupby(["問題類型", "來源檔案"]).size().reset_index(name="筆數")
    else:
        df_summary = pd.DataFrame(columns=["問題類型", "來源檔案", "筆數"])

    # ── 寫檔 ───────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("問題清單")
    write_df_to_sheet(ws1, df_issues if not df_issues.empty
                      else pd.DataFrame(columns=["問題類型", "來源檔案", "列號",
                                                  "欄位", "問題值", "問題描述"]))

    ws2 = wb.create_sheet("問題統計")
    write_df_to_sheet(ws2, df_summary, header_fill=FILL_ORANGE)

    ws3 = wb.create_sheet("清理日誌")
    write_df_to_sheet(ws3, df_log, header_fill=FILL_GREEN)

    wb.save(os.path.join(OUT_DIR, "data_quality_report.xlsx"))
    print("✔ data_quality_report.xlsx 完成（3 sheets）")

    return df_issues


# ══════════════════════════════════════════════════════════════
#  主程式
# ══════════════════════════════════════════════════════════════

def print_summary(df_sales: pd.DataFrame, df_budget: pd.DataFrame,
                  kpi: pd.DataFrame, profit: pd.DataFrame,
                  df_issues: pd.DataFrame):
    """輸出關鍵指標摘要。"""
    df = df_sales[
        (df_sales["銷售金額"] > 0) &
        (df_sales["數量"]     > 0)
    ].copy()
    df["銷售金額"] = pd.to_numeric(df["銷售金額"], errors="coerce")

    total_rev    = df["銷售金額"].sum()
    total_orders = len(df)
    total_sps    = df["業務員"].nunique()
    avg_order    = total_rev / total_orders if total_orders else 0
    top_sp       = (
        df.groupby("業務員")["銷售金額"].sum()
          .idxmax()
    ) if total_orders else "N/A"
    top_sp_amt   = df.groupby("業務員")["銷售金額"].sum().max()
    avg_achieve  = kpi["達成率%"].mean()
    over_target  = (kpi["達成率%"] >= 100).sum()
    total_sps_kpi = len(kpi)
    avg_margin   = profit["毛利率%"].mean()

    print("\n" + "═" * 58)
    print("  📊 關鍵指標摘要")
    print("═" * 58)
    print(f"  年度總銷售額    : NT$ {total_rev:>14,.0f}")
    print(f"  有效訂單數      : {total_orders:>10,} 筆")
    print(f"  業務員數        : {total_sps:>10} 人")
    print(f"  平均客單價      : NT$ {avg_order:>14,.0f}")
    print(f"  業績冠軍        : {top_sp}（NT$ {top_sp_amt:,.0f}）")
    print(f"  KPI 達成人數    : {over_target}/{total_sps_kpi} 人（平均達成率 {avg_achieve:.1f}%）")
    print(f"  平均產品毛利率  : {avg_margin:.1f}%")
    print(f"  資料品質問題    : {len(df_issues)} 筆")
    print("═" * 58)

    # 輸出檔案清單
    print("\n  📁 output/ 產出檔案：")
    for fn in sorted(os.listdir(OUT_DIR)):
        fpath = os.path.join(OUT_DIR, fn)
        size  = os.path.getsize(fpath)
        if size > 1024 * 1024:
            size_str = f"{size / 1024 / 1024:.1f} MB"
        else:
            size_str = f"{size / 1024:.1f} KB"
        print(f"    {fn:<45} {size_str:>8}")
    print()


def main():
    print("=" * 58)
    print("  03-advanced 多維度銷售分析報表產生器")
    print("=" * 58)
    print()

    print("【任務一】資料清理與標準化...")
    df_sales, df_budget, df_fb, df_log = clean_data()

    print("\n【任務二】多維度銷售分析...")
    df_sales_clean, df_merged = build_sales_analysis(df_sales)

    print("\n【任務三】預算達成率 KPI...")
    kpi = build_kpi_dashboard(df_sales, df_budget)

    print("\n【任務四】產品利潤交叉分析...")
    profit = build_product_profit(df_merged)

    print("\n【任務五】資料品質報告...")
    df_sales_raw  = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales.xlsx"))
    df_budget_raw = pd.read_excel(os.path.join(RAW_DIR, "budget_targets.xlsx"))
    df_fb_raw     = pd.read_excel(os.path.join(RAW_DIR, "customer_feedback.xlsx"))
    df_issues     = build_data_quality(df_sales_raw, df_budget_raw, df_fb_raw, df_log)

    print_summary(df_sales, df_budget, kpi, profit, df_issues)


if __name__ == "__main__":
    main()
