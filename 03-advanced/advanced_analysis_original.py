"""
進階教學 — 多維度銷售分析腳本
任務一：資料清理與標準化
任務二：多維度銷售分析（樞紐分析、趨勢、排名）
任務三：預算達成率分析與 KPI 儀表板
任務四：產品利潤分析與客戶滿意度交叉分析
任務五：產出精美 Excel 報表（含條件格式、圖表工作表）
"""

import re
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule, DataBarRule

RAW_DIR = Path(__file__).parent / "raw"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════
# 讀取原始資料
# ══════════════════════════════════════════════════════
sales_raw = pd.read_excel(RAW_DIR / "monthly_sales.xlsx", dtype=str)
budget_raw = pd.read_excel(RAW_DIR / "budget_targets.xlsx", dtype=str)
catalog_raw = pd.read_excel(RAW_DIR / "product_catalog.xlsx", dtype=str)
feedback_raw = pd.read_excel(RAW_DIR / "customer_feedback.xlsx", dtype=str)

sales_orig = sales_raw.copy()
budget_orig = budget_raw.copy()

# ══════════════════════════════════════════════════════
# 任務一：資料清理與標準化
# ══════════════════════════════════════════════════════
print("── 任務一：資料清理與標準化 ──")

cleaning_log = []


def log_clean(source, row_num, col, old_val, new_val, desc):
    cleaning_log.append({
        "來源檔案": source,
        "列號": row_num,
        "欄位": col,
        "原始值": str(old_val),
        "修正值": str(new_val),
        "問題描述": desc,
    })


# --- 清理銷售明細 ---
sales = sales_raw.copy()

# 1. 日期格式統一
for idx, val in sales["訂單日期"].items():
    if pd.notna(val) and "/" in str(val):
        new_val = str(val).replace("/", "-")
        log_clean("monthly_sales.xlsx", idx + 2, "訂單日期", val, new_val, "日期分隔符 / → -")
        sales.at[idx, "訂單日期"] = new_val

# 2. 業務員姓名去空白
for idx, val in sales["業務員"].items():
    if pd.notna(val):
        cleaned = str(val).replace(" ", "").strip()
        if cleaned != str(val):
            log_clean("monthly_sales.xlsx", idx + 2, "業務員", val, cleaned, "姓名含多餘空白")
            sales.at[idx, "業務員"] = cleaned

# 3. 產品名稱去空白與修正錯字
for idx, val in sales["產品名稱"].items():
    if pd.notna(val):
        cleaned = str(val).strip()
        if "智彗" in cleaned:
            cleaned = cleaned.replace("智彗", "智慧")
            log_clean("monthly_sales.xlsx", idx + 2, "產品名稱", val, cleaned, "產品名稱錯字修正")
        elif cleaned != str(val):
            log_clean("monthly_sales.xlsx", idx + 2, "產品名稱", val, cleaned, "產品名稱含多餘空白")
        sales.at[idx, "產品名稱"] = cleaned

# 4. 轉換數值欄位
for col in ["數量", "單價", "銷售金額"]:
    sales[col] = pd.to_numeric(sales[col], errors="coerce")

# --- 清理預算目標 ---
budget = budget_raw.copy()
for idx, val in budget["業務員"].items():
    if pd.notna(val):
        cleaned = re.sub(r"[（(].*?[）)]", "", str(val)).strip()
        if cleaned != str(val).strip():
            log_clean("budget_targets.xlsx", idx + 2, "業務員", val, cleaned, "業務員姓名含多餘標註")
            budget.at[idx, "業務員"] = cleaned
        elif str(val).strip() != str(val):
            log_clean("budget_targets.xlsx", idx + 2, "業務員", val, str(val).strip(), "姓名含多餘空白")
            budget.at[idx, "業務員"] = str(val).strip()

for col in ["年度目標金額", "Q1目標", "Q2目標", "Q3目標", "Q4目標"]:
    budget[col] = pd.to_numeric(budget[col], errors="coerce")

# --- 清理客戶回饋 ---
feedback = feedback_raw.copy()
for idx, val in feedback["日期"].items():
    if pd.notna(val) and "/" in str(val):
        new_val = str(val).replace("/", "-")
        log_clean("customer_feedback.xlsx", idx + 2, "日期", val, new_val, "日期分隔符 / → -")
        feedback.at[idx, "日期"] = new_val

feedback["滿意度評分"] = pd.to_numeric(feedback["滿意度評分"], errors="coerce")

# 修正產品名稱
for idx, val in feedback["產品名稱"].items():
    if pd.notna(val):
        cleaned = str(val).strip()
        cleaned = cleaned.replace("藍芽", "藍牙")
        if "機械鍵盤87" in cleaned:
            cleaned = cleaned.replace("機械鍵盤87鍵", "機械鍵盤 87鍵")
        if cleaned != str(val):
            log_clean("customer_feedback.xlsx", idx + 2, "產品名稱", val, cleaned, "產品名稱修正")
            feedback.at[idx, "產品名稱"] = cleaned

# --- 清理產品目錄 ---
catalog = catalog_raw.copy()
for col in ["建議售價", "成本", "庫存量"]:
    catalog[col] = pd.to_numeric(catalog[col], errors="coerce")

# 輸出清理後檔案
sales.to_excel(OUTPUT_DIR / "cleaned_monthly_sales.xlsx", index=False, engine="openpyxl")
budget.to_excel(OUTPUT_DIR / "cleaned_budget_targets.xlsx", index=False, engine="openpyxl")
feedback.to_excel(OUTPUT_DIR / "cleaned_customer_feedback.xlsx", index=False, engine="openpyxl")

# 輸出清理日誌
df_cleaning_log = pd.DataFrame(cleaning_log)
df_cleaning_log.to_excel(OUTPUT_DIR / "cleaning_log.xlsx", index=False, engine="openpyxl")

print(f"  ✔ 共修正 {len(cleaning_log)} 筆資料問題")
print("  ✔ 已輸出清理後檔案與清理日誌")


# ══════════════════════════════════════════════════════
# 任務二：多維度銷售分析
# ══════════════════════════════════════════════════════
print("\n── 任務二：多維度銷售分析 ──")

# 過濾有效銷售（金額 > 0 且數量 > 0）
valid_sales = sales[(sales["銷售金額"] > 0) & (sales["數量"] > 0)].copy()
valid_sales["月份"] = valid_sales["訂單日期"].str[:7]  # YYYY-MM

# 2-1：按月份的銷售趨勢
monthly_trend = valid_sales.groupby("月份").agg(
    訂單數=("訂單編號", "count"),
    銷售總額=("銷售金額", "sum"),
    平均客單價=("銷售金額", "mean"),
).reset_index()
monthly_trend["銷售總額"] = monthly_trend["銷售總額"].round(0).astype(int)
monthly_trend["平均客單價"] = monthly_trend["平均客單價"].round(0).astype(int)

# 計算月成長率 (MoM%)
monthly_trend["月成長率%"] = monthly_trend["銷售總額"].pct_change() * 100
monthly_trend["月成長率%"] = monthly_trend["月成長率%"].round(1)

# 2-2：按部門的銷售樞紐分析
dept_pivot = valid_sales.pivot_table(
    index="部門",
    columns="月份",
    values="銷售金額",
    aggfunc="sum",
    fill_value=0,
)
dept_pivot["年度合計"] = dept_pivot.sum(axis=1)
dept_pivot = dept_pivot.reset_index()

# 2-3：業務員排名（Top 10）
sales_ranking = valid_sales.groupby(["業務員", "部門"]).agg(
    訂單數=("訂單編號", "count"),
    銷售總額=("銷售金額", "sum"),
    平均每單金額=("銷售金額", "mean"),
).reset_index()
sales_ranking["銷售總額"] = sales_ranking["銷售總額"].round(0).astype(int)
sales_ranking["平均每單金額"] = sales_ranking["平均每單金額"].round(0).astype(int)
sales_ranking = sales_ranking.sort_values("銷售總額", ascending=False).reset_index(drop=True)
sales_ranking.index = sales_ranking.index + 1
sales_ranking.index.name = "排名"
sales_ranking = sales_ranking.reset_index()

# 2-4：按產品類別的銷售分布
category_sales = valid_sales.merge(
    catalog[["產品編號", "產品類別"]], on="產品編號", how="left"
)
category_summary = category_sales.groupby("產品類別").agg(
    銷售筆數=("訂單編號", "count"),
    銷售總額=("銷售金額", "sum"),
    銷售佔比=("銷售金額", lambda x: 0),  # placeholder
).reset_index()
total_sales = category_summary["銷售總額"].sum()
category_summary["銷售佔比"] = (category_summary["銷售總額"] / total_sales * 100).round(1)
category_summary["銷售佔比"] = category_summary["銷售佔比"].astype(str) + "%"
category_summary = category_summary.sort_values("銷售總額", ascending=False).reset_index(drop=True)

# 2-5：按客戶區域的銷售分布
region_summary = valid_sales.groupby("客戶區域").agg(
    銷售筆數=("訂單編號", "count"),
    銷售總額=("銷售金額", "sum"),
).reset_index().sort_values("銷售總額", ascending=False).reset_index(drop=True)

print("  ✔ 月度趨勢、部門樞紐、業務排名、產品類別、區域分布 分析完成")


# ══════════════════════════════════════════════════════
# 任務三：預算達成率分析與 KPI 儀表板
# ══════════════════════════════════════════════════════
print("\n── 任務三：預算達成率分析 ──")

# 計算每位業務的實際銷售額
actual_by_person = valid_sales.groupby("業務員")["銷售金額"].sum().reset_index()
actual_by_person.columns = ["業務員", "實際銷售額"]
actual_by_person["實際銷售額"] = actual_by_person["實際銷售額"].round(0).astype(int)

# 合併預算目標
kpi = budget[["業務員", "部門", "年度目標金額"]].merge(actual_by_person, on="業務員", how="left")
kpi["實際銷售額"] = kpi["實際銷售額"].fillna(0).astype(int)
kpi["年度目標金額"] = kpi["年度目標金額"].astype(int)
kpi["達成率%"] = (kpi["實際銷售額"] / kpi["年度目標金額"] * 100).round(1)
kpi["差異"] = kpi["實際銷售額"] - kpi["年度目標金額"]

# 達成狀態
def get_status(rate):
    if rate >= 120:
        return "★ 超標達成"
    elif rate >= 100:
        return "✔ 達成"
    elif rate >= 80:
        return "△ 接近達成"
    else:
        return "✘ 未達成"

kpi["達成狀態"] = kpi["達成率%"].apply(get_status)
kpi = kpi.sort_values("達成率%", ascending=False).reset_index(drop=True)

# 部門 KPI 彙總
dept_kpi = kpi.groupby("部門").agg(
    業務人數=("業務員", "count"),
    目標合計=("年度目標金額", "sum"),
    實際合計=("實際銷售額", "sum"),
).reset_index()
dept_kpi["部門達成率%"] = (dept_kpi["實際合計"] / dept_kpi["目標合計"] * 100).round(1)
dept_kpi = dept_kpi.sort_values("部門達成率%", ascending=False).reset_index(drop=True)

print(f"  ✔ 個人 KPI {len(kpi)} 筆、部門 KPI {len(dept_kpi)} 筆")


# ══════════════════════════════════════════════════════
# 任務四：產品利潤分析與客戶滿意度交叉分析
# ══════════════════════════════════════════════════════
print("\n── 任務四：產品利潤與滿意度交叉分析 ──")

# 產品銷售彙總
product_sales = valid_sales.groupby("產品編號").agg(
    銷售數量=("數量", "sum"),
    銷售總額=("銷售金額", "sum"),
).reset_index()

# 合併產品目錄取得成本
product_analysis = product_sales.merge(
    catalog[["產品編號", "產品名稱", "產品類別", "建議售價", "成本"]],
    on="產品編號", how="left"
)
product_analysis["估計成本總額"] = product_analysis["銷售數量"] * product_analysis["成本"]
product_analysis["估計毛利"] = product_analysis["銷售總額"] - product_analysis["估計成本總額"]
product_analysis["毛利率%"] = (
    product_analysis["估計毛利"] / product_analysis["銷售總額"] * 100
).round(1)

# 合併滿意度
valid_feedback = feedback[
    (feedback["滿意度評分"] >= 1) & (feedback["滿意度評分"] <= 5)
].copy()
avg_score = valid_feedback.groupby("產品編號")["滿意度評分"].agg(
    平均滿意度="mean",
    回饋筆數="count",
).reset_index()
avg_score["平均滿意度"] = avg_score["平均滿意度"].round(2)

product_analysis = product_analysis.merge(avg_score, on="產品編號", how="left")
product_analysis = product_analysis.sort_values("估計毛利", ascending=False).reset_index(drop=True)

# 整理欄位順序
product_analysis = product_analysis[[
    "產品編號", "產品名稱", "產品類別",
    "銷售數量", "銷售總額", "建議售價", "成本",
    "估計成本總額", "估計毛利", "毛利率%",
    "平均滿意度", "回饋筆數",
]]

print(f"  ✔ 產品利潤分析 {len(product_analysis)} 項產品")


# ══════════════════════════════════════════════════════
# 任務五：資料品質報告
# ══════════════════════════════════════════════════════
print("\n── 任務五：資料品質報告 ──")

quality_issues = []

# 銷售明細問題
# 負數金額
neg_sales = sales_orig[pd.to_numeric(sales_orig["銷售金額"], errors="coerce") < 0]
for idx, row in neg_sales.iterrows():
    quality_issues.append({
        "來源檔案": "monthly_sales.xlsx",
        "列號": idx + 2,
        "問題類型": "邏輯異常",
        "欄位": "銷售金額",
        "原始值": row["銷售金額"],
        "問題描述": "銷售金額為負數（疑似退貨未標記）",
    })

# 數量為 0
zero_qty = sales_orig[pd.to_numeric(sales_orig["數量"], errors="coerce") == 0]
for idx, row in zero_qty.iterrows():
    quality_issues.append({
        "來源檔案": "monthly_sales.xlsx",
        "列號": idx + 2,
        "問題類型": "邏輯異常",
        "欄位": "數量",
        "原始值": row["數量"],
        "問題描述": "銷售數量為 0",
    })

# 重複訂單編號
dup_orders = sales_orig[sales_orig.duplicated(subset=["訂單編號"], keep=False)]
for idx, row in dup_orders.iterrows():
    quality_issues.append({
        "來源檔案": "monthly_sales.xlsx",
        "列號": idx + 2,
        "問題類型": "重複資料",
        "欄位": "訂單編號",
        "原始值": row["訂單編號"],
        "問題描述": "訂單編號重複出現",
    })

# 回饋評分超出範圍
invalid_scores = feedback_raw.copy()
invalid_scores["滿意度評分"] = pd.to_numeric(invalid_scores["滿意度評分"], errors="coerce")
out_of_range = invalid_scores[
    (invalid_scores["滿意度評分"] < 1) | (invalid_scores["滿意度評分"] > 5)
]
for idx, row in out_of_range.iterrows():
    quality_issues.append({
        "來源檔案": "customer_feedback.xlsx",
        "列號": idx + 2,
        "問題類型": "格式異常",
        "欄位": "滿意度評分",
        "原始值": str(row["滿意度評分"]),
        "問題描述": f"評分超出 1-5 有效範圍（值={row['滿意度評分']}）",
    })

# 預算 Q1-Q4 加總 vs 年度目標
for idx, row in budget.iterrows():
    q_sum = row["Q1目標"] + row["Q2目標"] + row["Q3目標"] + row["Q4目標"]
    annual = row["年度目標金額"]
    if annual > 0 and q_sum > annual * 1.15:
        quality_issues.append({
            "來源檔案": "budget_targets.xlsx",
            "列號": idx + 2,
            "問題類型": "邏輯異常",
            "欄位": "Q1-Q4目標",
            "原始值": f"Q合計={int(q_sum)}, 年度={int(annual)}",
            "問題描述": f"季度目標加總（{int(q_sum)}）超過年度目標的 115%",
        })

df_quality = pd.DataFrame(quality_issues)
print(f"  ✔ 共偵測到 {len(quality_issues)} 筆資料品質問題")


# ══════════════════════════════════════════════════════
# 輸出精美 Excel 報表
# ══════════════════════════════════════════════════════
print("\n── 輸出精美 Excel 報表 ──")

# ── 樣式定義 ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def style_header(ws, header_fill=HEADER_FILL):
    """為工作表的第一列套用標題樣式"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cells(ws):
    """為資料儲存格套用邊框"""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def auto_column_width(ws, min_width=10, max_width=30):
    """自動調整欄寬"""
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value:
                # 中文字算 2 個字元寬度
                val_str = str(cell.value)
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, char_len)
        adjusted = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted


def write_df_to_sheet(wb, df, sheet_name, header_fill=HEADER_FILL):
    """將 DataFrame 寫入帶格式的工作表"""
    ws = wb.create_sheet(title=sheet_name)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    style_header(ws, header_fill)
    style_data_cells(ws)
    auto_column_width(ws)
    return ws


# ═══════ 報表一：sales_analysis_report.xlsx ═══════
wb = load_workbook(OUTPUT_DIR / "cleaned_monthly_sales.xlsx") if False else None
from openpyxl import Workbook
wb = Workbook()
wb.remove(wb.active)  # 移除預設空白工作表

# Sheet 1: 月度趨勢
ws_trend = write_df_to_sheet(wb, monthly_trend, "月度銷售趨勢")

# 加入折線圖 — 銷售總額趨勢
chart_line = LineChart()
chart_line.title = "2024 年月度銷售趨勢"
chart_line.y_axis.title = "銷售金額"
chart_line.x_axis.title = "月份"
chart_line.style = 10
chart_line.width = 25
chart_line.height = 14

data_ref = Reference(ws_trend, min_col=3, min_row=1, max_row=ws_trend.max_row)
cats_ref = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)
chart_line.add_data(data_ref, titles_from_data=True)
chart_line.set_categories(cats_ref)
chart_line.series[0].graphicalProperties.line.width = 25000

ws_trend.add_chart(chart_line, "A15")

# 加入長條圖 — 訂單數
chart_bar = BarChart()
chart_bar.title = "月度訂單數量"
chart_bar.y_axis.title = "訂單數"
chart_bar.style = 10
chart_bar.width = 25
chart_bar.height = 14

data_ref2 = Reference(ws_trend, min_col=2, min_row=1, max_row=ws_trend.max_row)
chart_bar.add_data(data_ref2, titles_from_data=True)
chart_bar.set_categories(cats_ref)

ws_trend.add_chart(chart_bar, "A32")

# Sheet 2: 部門樞紐分析
ws_dept = write_df_to_sheet(wb, dept_pivot, "部門銷售樞紐", HEADER_FILL_GREEN)

# Sheet 3: 業務員排名
ws_rank = write_df_to_sheet(wb, sales_ranking, "業務員排名")

# 為 Top 3 加底色
for row_idx in range(2, min(5, ws_rank.max_row + 1)):
    for cell in ws_rank[row_idx]:
        cell.fill = GOOD_FILL

# Sheet 4: 產品類別分布
ws_cat = write_df_to_sheet(wb, category_summary, "產品類別分布", HEADER_FILL_ORANGE)

# 加入圓餅圖
pie_chart = PieChart()
pie_chart.title = "產品類別銷售佔比"
pie_chart.style = 10
pie_chart.width = 18
pie_chart.height = 14

labels = Reference(ws_cat, min_col=1, min_row=2, max_row=ws_cat.max_row)
data_pie = Reference(ws_cat, min_col=3, min_row=1, max_row=ws_cat.max_row)
pie_chart.add_data(data_pie, titles_from_data=True)
pie_chart.set_categories(labels)
ws_cat.add_chart(pie_chart, "A10")

# Sheet 5: 區域分布
write_df_to_sheet(wb, region_summary, "客戶區域分布")

wb.save(OUTPUT_DIR / "sales_analysis_report.xlsx")
print("  ✔ 已輸出 sales_analysis_report.xlsx（含圖表）")


# ═══════ 報表二：kpi_dashboard.xlsx ═══════
wb2 = Workbook()
wb2.remove(wb2.active)

# Sheet 1: KPI 總覽
ws_kpi = write_df_to_sheet(wb2, kpi, "個人KPI達成率")

# 為達成率加條件格式
rate_col = None
for col_idx, cell in enumerate(ws_kpi[1], 1):
    if cell.value == "達成率%":
        rate_col = col_idx
        break

if rate_col:
    col_letter = ws_kpi.cell(row=1, column=rate_col).column_letter
    range_str = f"{col_letter}2:{col_letter}{ws_kpi.max_row}"

    ws_kpi.conditional_formatting.add(
        range_str,
        CellIsRule(operator="greaterThanOrEqual", formula=["100"], fill=GOOD_FILL)
    )
    ws_kpi.conditional_formatting.add(
        range_str,
        CellIsRule(operator="between", formula=["80", "99.9"], fill=WARN_FILL)
    )
    ws_kpi.conditional_formatting.add(
        range_str,
        CellIsRule(operator="lessThan", formula=["80"], fill=BAD_FILL)
    )

# Sheet 2: 部門 KPI
ws_dept_kpi = write_df_to_sheet(wb2, dept_kpi, "部門KPI彙總", HEADER_FILL_GREEN)

# 加入長條圖
chart_kpi = BarChart()
chart_kpi.title = "各部門年度達成率"
chart_kpi.y_axis.title = "達成率 %"
chart_kpi.style = 10
chart_kpi.width = 20
chart_kpi.height = 14

kpi_data = Reference(ws_dept_kpi, min_col=5, min_row=1, max_row=ws_dept_kpi.max_row)
kpi_cats = Reference(ws_dept_kpi, min_col=1, min_row=2, max_row=ws_dept_kpi.max_row)
chart_kpi.add_data(kpi_data, titles_from_data=True)
chart_kpi.set_categories(kpi_cats)

ws_dept_kpi.add_chart(chart_kpi, "A10")

wb2.save(OUTPUT_DIR / "kpi_dashboard.xlsx")
print("  ✔ 已輸出 kpi_dashboard.xlsx（含條件格式與圖表）")


# ═══════ 報表三：product_profit_report.xlsx ═══════
wb3 = Workbook()
wb3.remove(wb3.active)

ws_profit = write_df_to_sheet(wb3, product_analysis, "產品利潤與滿意度")

# 為毛利率加資料條
margin_col = None
for col_idx, cell in enumerate(ws_profit[1], 1):
    if cell.value == "毛利率%":
        margin_col = col_idx
        break

if margin_col:
    col_letter = ws_profit.cell(row=1, column=margin_col).column_letter
    range_str = f"{col_letter}2:{col_letter}{ws_profit.max_row}"
    ws_profit.conditional_formatting.add(
        range_str,
        DataBarRule(start_type="min", end_type="max", color="5B9BD5")
    )

wb3.save(OUTPUT_DIR / "product_profit_report.xlsx")
print("  ✔ 已輸出 product_profit_report.xlsx（含資料條格式）")


# ═══════ 報表四：data_quality_report.xlsx ═══════
wb4 = Workbook()
wb4.remove(wb4.active)

# Sheet 1: 問題清單
ws_quality = write_df_to_sheet(wb4, df_quality, "資料品質問題清單")

# Sheet 2: 問題統計
issue_summary = df_quality.groupby(["問題類型", "來源檔案"]).size().reset_index(name="筆數")
ws_issue_sum = write_df_to_sheet(wb4, issue_summary, "問題統計", HEADER_FILL_ORANGE)

# Sheet 3: 清理日誌
write_df_to_sheet(wb4, df_cleaning_log, "清理修正日誌", HEADER_FILL_GREEN)

wb4.save(OUTPUT_DIR / "data_quality_report.xlsx")
print("  ✔ 已輸出 data_quality_report.xlsx")


# ══════════════════════════════════════════════════════
# 完成摘要
# ══════════════════════════════════════════════════════
print("\n" + "═" * 50)
print("全部處理完成！產出檔案清單：")
print("═" * 50)
for f in sorted(OUTPUT_DIR.iterdir()):
    size_kb = f.stat().st_size / 1024
    print(f"  {f.name:45s} ({size_kb:.1f} KB)")

print(f"\n共 {len(list(OUTPUT_DIR.iterdir()))} 個檔案於 03-advanced/output/")

# 顯示關鍵指標摘要
print("\n── 關鍵指標摘要 ──")
print(f"  有效銷售筆數：{len(valid_sales)}")
print(f"  年度銷售總額：{valid_sales['銷售金額'].sum():,.0f} 元")
print(f"  業務員人數：  {valid_sales['業務員'].nunique()}")
print(f"  產品項目數：  {valid_sales['產品編號'].nunique()}")
print(f"  資料品質問題：{len(quality_issues)} 筆")
print(f"  格式修正紀錄：{len(cleaning_log)} 筆")
