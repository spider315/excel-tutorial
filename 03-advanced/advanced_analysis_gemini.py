import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
import re

# 設定路徑
RAW_DIR = '03-advanced/raw'
OUTPUT_DIR = '03-advanced/output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 顏色定義
COLOR_BLUE = '2F5496'
COLOR_GREEN = '548235'
COLOR_ORANGE = 'BF8F00'
COLOR_LIGHT_GREEN = 'C6EFCE'
COLOR_LIGHT_YELLOW = 'FFEB9C'
COLOR_LIGHT_RED = 'FFC7CE'
COLOR_WHITE = 'FFFFFF'
COLOR_DATA_BAR = '5B9BD5'

# 邊框定義
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_excel(ws, title_color=COLOR_BLUE):
    """通用格式化函數"""
    # 標題列格式
    header_font = Font(bold=True, color=COLOR_WHITE)
    header_fill = PatternFill(start_color=title_color, end_color=title_color, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = THIN_BORDER

    # 內容列格式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center')
            cell.border = THIN_BORDER

    # 自動欄寬
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                # 估計寬度 (中文算2字)
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_length:
                    max_length = length
            except:
                pass
        adjusted_width = min(max(max_length + 2, 10), 30)
        ws.column_dimensions[column].width = adjusted_width

def add_data_bar(ws, col_index, start_row, end_row, color):
    """模擬數據條 (openpyxl 不直接支援 Data Bar 規則但可設底色，這裡僅做簡化處理或使用條件格式)"""
    from openpyxl.formatting.rule import DataBarRule
    col_letter = get_column_letter(col_index)
    ref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    rule = DataBarRule(start_type='min', end_type='max', color=color)
    ws.conditional_formatting.add(ref, rule)

# --- 任務一：資料清理與標準化 ---
print("執行任務一：資料清理...")
cleaning_logs = []

def log_fix(source, row, col, old, new, desc):
    cleaning_logs.append({
        '來源檔案': source,
        '列號': row + 2, # Pandas index to Excel row
        '欄位': col,
        '原始值': old,
        '修正值': new,
        '問題描述': desc
    })

# 1. 讀取資料
df_sales = pd.read_excel(os.path.join(RAW_DIR, 'monthly_sales.xlsx'))
df_budget = pd.read_excel(os.path.join(RAW_DIR, 'budget_targets.xlsx'))
df_feedback = pd.read_excel(os.path.join(RAW_DIR, 'customer_feedback.xlsx'))
df_catalog = pd.read_excel(os.path.join(RAW_DIR, 'product_catalog.xlsx'))

# 清理規則實作
def clean_df(df, filename):
    # A. 日期格式統一 (YYYY-MM-DD)
    date_cols = [c for c in df.columns if '日期' in c or '時間' in c or c == '交貨期']
    for col in date_cols:
        for idx, val in df[col].items():
            original = str(val)
            new_val = pd.to_datetime(val).strftime('%Y-%m-%d')
            if original != new_val:
                df.at[idx, col] = new_val
                log_fix(filename, idx, col, original, new_val, "日期格式不統一")

    # B. 業務員姓名清理
    if '業務員' in df.columns:
        for idx, val in df['業務員'].items():
            original = str(val)
            # 去空白
            new_val = original.strip()
            # 移除括號標註
            new_val = re.sub(r'[\(（].*?[\)）]', '', new_val)
            if original != new_val:
                df.at[idx, '業務員'] = new_val
                log_fix(filename, idx, '業務員', original, new_val, "姓名含空格或括號")

    # C. 產品名稱清理
    if '產品名稱' in df.columns:
        for idx, val in df['產品名稱'].items():
            original = str(val)
            new_val = original.strip()
            # 修正錯字
            new_val = new_val.replace('智彗', '智慧').replace('藍芽', '藍牙')
            if original != new_val:
                df.at[idx, '產品名稱'] = new_val
                log_fix(filename, idx, '產品名稱', original, new_val, "名稱錯字或空格")

    # D. 數值型態轉換
    numeric_keywords = ['金額', '單價', '數量', '預算', '評分', '成本', 'Q1', 'Q2', 'Q3', 'Q4']
    for col in df.columns:
        if any(k in col for k in numeric_keywords):
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    return df

df_sales = clean_df(df_sales, 'monthly_sales.xlsx')
df_budget = clean_df(df_budget, 'budget_targets.xlsx')
df_feedback = clean_df(df_feedback, 'customer_feedback.xlsx')

# 儲存清理後資料
df_sales.to_excel(os.path.join(OUTPUT_DIR, 'cleaned_monthly_sales.xlsx'), index=False)
df_budget.to_excel(os.path.join(OUTPUT_DIR, 'cleaned_budget_targets.xlsx'), index=False)
df_feedback.to_excel(os.path.join(OUTPUT_DIR, 'cleaned_customer_feedback.xlsx'), index=False)

df_logs = pd.DataFrame(cleaning_logs)
df_logs.to_excel(os.path.join(OUTPUT_DIR, 'cleaning_log.xlsx'), index=False)

# 格式化清理後的檔案
for f in ['cleaned_monthly_sales.xlsx', 'cleaned_budget_targets.xlsx', 'cleaned_customer_feedback.xlsx', 'cleaning_log.xlsx']:
    path = os.path.join(OUTPUT_DIR, f)
    wb = Workbook()
    ws = wb.active
    # Reload cleaned data into openpyxl for formatting
    temp_df = pd.read_excel(path)
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(temp_df, index=False, header=True):
        ws.append(r)
    format_excel(ws)
    wb.save(path)

# --- 任務二：多維度銷售分析 ---
print("執行任務二：銷售分析...")
# 排除銷售金額 <= 0 或數量 <= 0
df_sales_valid = df_sales[(df_sales['銷售金額'] > 0) & (df_sales['數量'] > 0)].copy()

wb_sales = Workbook()
# Sheet 1: 月度銷售趨勢
ws1 = wb_sales.active
ws1.title = "月度銷售趨勢"
df_sales_valid['月份'] = pd.to_datetime(df_sales_valid['訂單日期']).dt.strftime('%Y-%m')
trend = df_sales_valid.groupby('月份').agg(
    訂單數=('訂單編號', 'count'),
    銷售總額=('銷售金額', 'sum')
).reset_index()
trend['平均客單價'] = (trend['銷售總額'] / trend['訂單數']).round(2)
trend['月成長率%'] = trend['銷售總額'].pct_change().fillna(0) * 100
trend['月成長率%'] = trend['月成長率%'].round(2)

for r in dataframe_to_rows(trend, index=False, header=True):
    ws1.append(r)
format_excel(ws1)

# 加入趨勢圖表
# 銷售總額折線圖 (A15)
lc = LineChart()
lc.title = "月度銷售總額趨勢"
lc.y_axis.title = "金額"
lc.x_axis.title = "月份"
data = Reference(ws1, min_col=3, min_row=1, max_row=len(trend)+1)
cats = Reference(ws1, min_col=1, min_row=2, max_row=len(trend)+1)
lc.add_data(data, titles_from_data=True)
lc.set_categories(cats)
ws1.add_chart(lc, "A15")

# 訂單數長條圖 (A32)
bc = BarChart()
bc.title = "月度訂單數"
data = Reference(ws1, min_col=2, min_row=1, max_row=len(trend)+1)
bc.add_data(data, titles_from_data=True)
bc.set_categories(cats)
ws1.add_chart(bc, "A32")

# Sheet 2: 部門銷售樞紐
ws2 = wb_sales.create_sheet("部門銷售樞紐")
pivot_dept = df_sales_valid.pivot_table(
    index='部門', columns='月份', values='銷售金額', aggfunc='sum', fill_value=0
)
pivot_dept['年度合計'] = pivot_dept.sum(axis=1)
pivot_dept.reset_index(inplace=True)

for r in dataframe_to_rows(pivot_dept, index=False, header=True):
    ws2.append(r)
format_excel(ws2)

# Sheet 3: 業務員排名
ws3 = wb_sales.create_sheet("業務員排名")
ranking = df_sales_valid.groupby(['業務員', '部門']).agg(
    訂單數=('訂單編號', 'count'),
    銷售總額=('銷售金額', 'sum')
).reset_index()
ranking['平均每單金額'] = (ranking['銷售總額'] / ranking['訂單數']).round(2)
ranking = ranking.sort_values('銷售總額', ascending=False).reset_index(drop=True)
ranking.insert(0, '排名', ranking.index + 1)

for r in dataframe_to_rows(ranking, index=False, header=True):
    ws3.append(r)
format_excel(ws3)

# Top 3 加淺綠底
green_fill = PatternFill(start_color=COLOR_LIGHT_GREEN, end_color=COLOR_LIGHT_GREEN, fill_type='solid')
for r in range(2, 5): # Header is row 1, Top 3 are 2,3,4
    if r <= ws3.max_row:
        for cell in ws3[r]:
            cell.fill = green_fill

# Sheet 4: 產品類別分布
ws4 = wb_sales.create_sheet("產品類別分布")
df_merged_prod = df_sales_valid.merge(df_catalog[['產品編號', '產品類別']], on='產品編號', how='left')
prod_dist = df_merged_prod.groupby('產品類別').agg(
    銷售筆數=('訂單編號', 'count'),
    銷售總額=('銷售金額', 'sum')
).reset_index()
total_sales = prod_dist['銷售總額'].sum()
prod_dist['銷售佔比'] = (prod_dist['銷售總額'] / total_sales).round(4)

for r in dataframe_to_rows(prod_dist, index=False, header=True):
    ws4.append(r)
format_excel(ws4)

# 圓餅圖
pc = PieChart()
pc.title = "產品類別銷售佔比"
data = Reference(ws4, min_col=3, min_row=2, max_row=len(prod_dist)+1)
cats = Reference(ws4, min_col=1, min_row=2, max_row=len(prod_dist)+1)
pc.add_data(data)
pc.set_categories(cats)
ws4.add_chart(pc, "A10")

# Sheet 5: 客戶區域分布
ws5 = wb_sales.create_sheet("客戶區域分布")
region_dist = df_sales_valid.groupby('客戶區域').agg(
    銷售筆數=('訂單編號', 'count'),
    銷售總額=('銷售金額', 'sum')
).reset_index().sort_values('銷售總額', ascending=False)

for r in dataframe_to_rows(region_dist, index=False, header=True):
    ws5.append(r)
format_excel(ws5)

wb_sales.save(os.path.join(OUTPUT_DIR, 'sales_analysis_report.xlsx'))

# --- 任務三：預算達成率 ---
print("執行任務三：KPI 儀表板...")
wb_kpi = Workbook()
# Sheet 1: 個人KPI達成率
ws_kpi1 = wb_kpi.active
ws_kpi1.title = "個人KPI達成率"

# 尋找欄位
col_salesperson = next(c for c in df_budget.columns if '業務員' in c)
col_dept = next(c for c in df_budget.columns if '部門' in c)
col_budget_total = next(c for c in df_budget.columns if '年度預算' in c or '年度' in c)

actual_sales = df_sales_valid.groupby('業務員')['銷售金額'].sum().reset_index()
kpi_data = actual_sales.merge(df_budget[[col_salesperson, col_dept, col_budget_total]], left_on='業務員', right_on=col_salesperson, how='left')
# 重新命名欄位以便後續操作
kpi_data = kpi_data.drop(columns=[col_salesperson])
kpi_data = kpi_data.rename(columns={
    '業務員': '業務員',
    '銷售金額': '實際銷售額',
    col_dept: '部門',
    col_budget_total: '預算目標'
})
kpi_data['達成率%'] = (kpi_data['實際銷售額'] / kpi_data['預算目標'] * 100).round(2)

def get_status(row):
    rate = row['達成率%']
    if rate >= 120: return '★超標'
    if rate >= 100: return '✔達成'
    if rate >= 80: return '△接近'
    return '✘未達成'

kpi_data['狀態'] = kpi_data.apply(get_status, axis=1)

for r in dataframe_to_rows(kpi_data, index=False, header=True):
    ws_kpi1.append(r)
format_excel(ws_kpi1, COLOR_GREEN)

# 條件格式
from openpyxl.formatting.rule import CellIsRule
rate_col = get_column_letter(kpi_data.columns.get_loc('達成率%') + 1)
rows = len(kpi_data) + 1
ws_kpi1.conditional_formatting.add(f'{rate_col}2:{rate_col}{rows}',
    CellIsRule(operator='greaterThanOrEqual', formula=['100'], fill=PatternFill(start_color=COLOR_LIGHT_GREEN, end_color=COLOR_LIGHT_GREEN, fill_type='solid')))
ws_kpi1.conditional_formatting.add(f'{rate_col}2:{rate_col}{rows}',
    CellIsRule(operator='between', formula=['80', '99.99'], fill=PatternFill(start_color=COLOR_LIGHT_YELLOW, end_color=COLOR_LIGHT_YELLOW, fill_type='solid')))
ws_kpi1.conditional_formatting.add(f'{rate_col}2:{rate_col}{rows}',
    CellIsRule(operator='lessThan', formula=['80'], fill=PatternFill(start_color=COLOR_LIGHT_RED, end_color=COLOR_LIGHT_RED, fill_type='solid')))

# Sheet 2: 部門KPI彙總
ws_kpi2 = wb_kpi.create_sheet("部門KPI彙總")
dept_kpi = kpi_data.groupby('部門').agg({
    '實際銷售額': 'sum',
    '預算目標': 'sum'
}).reset_index()
dept_kpi['達成率%'] = (dept_kpi['實際銷售額'] / dept_kpi['預算目標'] * 100).round(2)

for r in dataframe_to_rows(dept_kpi, index=False, header=True):
    ws_kpi2.append(r)
format_excel(ws_kpi2, COLOR_GREEN)

# 長條圖
bc_kpi = BarChart()
bc_kpi.title = "部門預算達成狀況"
data = Reference(ws_kpi2, min_col=2, max_col=3, min_row=1, max_row=len(dept_kpi)+1)
cats = Reference(ws_kpi2, min_col=1, min_row=2, max_row=len(dept_kpi)+1)
bc_kpi.add_data(data, titles_from_data=True)
bc_kpi.set_categories(cats)
ws_kpi2.add_chart(bc_kpi, "A10")

wb_kpi.save(os.path.join(OUTPUT_DIR, 'kpi_dashboard.xlsx'))

# --- 任務四：產品利潤交叉分析 ---
print("執行任務四：產品利潤分析...")
wb_profit = Workbook()
ws_p = wb_profit.active
ws_p.title = "產品利潤分析"

prod_sales = df_sales_valid.groupby('產品編號').agg(
    銷售總額=('銷售金額', 'sum'),
    數量=('數量', 'sum')
).reset_index()

prod_profit = prod_sales.merge(df_catalog[['產品編號', '產品名稱', '產品類別', '成本']], on='產品編號', how='left')
avg_feedback = df_feedback.groupby('產品編號')['滿意度評分'].mean().reset_index()
prod_profit = prod_profit.merge(avg_feedback, on='產品編號', how='left')

prod_profit['總成本'] = prod_profit['數量'] * prod_profit['成本']
prod_profit['毛利'] = prod_profit['銷售總額'] - prod_profit['總成本']
prod_profit['毛利率%'] = (prod_profit['毛利'] / prod_profit['銷售總額'] * 100).round(2)
prod_profit = prod_profit.sort_values('毛利', ascending=False).fillna(0)

# 重整欄位順序
cols = ['產品編號', '產品名稱', '產品類別', '數量', '銷售總額', '毛利', '毛利率%', '滿意度評分']
prod_profit = prod_profit[cols]

for r in dataframe_to_rows(prod_profit, index=False, header=True):
    ws_p.append(r)
format_excel(ws_p, COLOR_ORANGE)

# 資料條
margin_col_idx = cols.index('毛利率%') + 1
add_data_bar(ws_p, margin_col_idx, 2, len(prod_profit)+1, COLOR_DATA_BAR)

wb_profit.save(os.path.join(OUTPUT_DIR, 'product_profit_report.xlsx'))

# --- 任務五：資料品質報告 ---
print("執行任務五：資料品質報告...")
problems = []

# 負數金額或數量0 (來自原始 df_sales)
invalid_sales = df_sales[(df_sales['銷售金額'] <= 0) | (df_sales['數量'] <= 0)]
for idx, row in invalid_sales.iterrows():
    problems.append({
        '問題類型': '負數金額或數量0',
        '來源檔案': 'monthly_sales.xlsx',
        '詳細資訊': f"訂單 {row['訂單編號']}, 金額 {row['銷售金額']}, 數量 {row['數量']}"
    })

# 重複訂單
dupes = df_sales[df_sales.duplicated('訂單編號', keep=False)]
for idx, row in dupes.iterrows():
    problems.append({
        '問題類型': '重複訂單',
        '來源檔案': 'monthly_sales.xlsx',
        '詳細資訊': f"訂單編號 {row['訂單編號']} 重複"
    })

# 評分超範圍 (1-5)
out_of_range_fb = df_feedback[(df_feedback['滿意度評分'] < 1) | (df_feedback['滿意度評分'] > 5)]
for idx, row in out_of_range_fb.iterrows():
    problems.append({
        '問題類型': '評分超範圍',
        '來源檔案': 'customer_feedback.xlsx',
        '詳細資訊': f"回饋 {row['回饋編號']}, 評分 {row['滿意度評分']}"
    })

# 預算Q合計超115%
q_cols = [c for c in df_budget.columns if any(q in c for q in ['Q1', 'Q2', 'Q3', 'Q4'])]
if q_cols:
    df_budget['Q合計'] = df_budget[q_cols].sum(axis=1)
    budget_error = df_budget[df_budget['Q合計'] > df_budget[col_budget_total] * 1.15]
    for idx, row in budget_error.iterrows():
        problems.append({
            '問題類型': '預算Q合計超115%',
            '來源檔案': 'budget_targets.xlsx',
            '詳細資訊': f"業務員 {row[col_salesperson]}, Q合計 {row['Q合計']}, 年度總額 {row[col_budget_total]}"
        })

wb_qa = Workbook()
# Sheet 1: 問題清單
ws_qa1 = wb_qa.active
ws_qa1.title = "問題清單"
df_qa1 = pd.DataFrame(problems)
for r in dataframe_to_rows(df_qa1, index=False, header=True):
    ws_qa1.append(r)
format_excel(ws_qa1, 'C00000')

# Sheet 2: 問題統計
ws_qa2 = wb_qa.create_sheet("問題統計")
if not df_qa1.empty:
    qa_stats = df_qa1.groupby(['問題類型', '來源檔案']).size().reset_index(name='個數')
    for r in dataframe_to_rows(qa_stats, index=False, header=True):
        ws_qa2.append(r)
    format_excel(ws_qa2, 'C00000')

# Sheet 3: 清理日誌
ws_qa3 = wb_qa.create_sheet("清理日誌")
for r in dataframe_to_rows(df_logs, index=False, header=True):
    ws_qa3.append(r)
format_excel(ws_qa3, 'C00000')

wb_qa.save(os.path.join(OUTPUT_DIR, 'data_quality_report.xlsx'))

print("\n分析完成！")

# 輸出檔案資訊
print("\n--- 輸出檔案清單 ---")
for f in os.listdir(OUTPUT_DIR):
    path = os.path.join(OUTPUT_DIR, f)
    size = os.path.getsize(path) / 1024
    print(f"{f:<30} {size:>8.2f} KB")

# 關鍵指標摘要
total_rev = df_sales_valid['銷售金額'].sum()
avg_margin = prod_profit['毛利率%'].mean()
problem_count = len(problems)
print(f"\n--- 關鍵指標摘要 ---")
print(f"總銷售額: {total_rev:,.0f}")
print(f"平均產品毛利率: {avg_margin:.2f}%")
print(f"發現數據品質問題數: {problem_count}")
