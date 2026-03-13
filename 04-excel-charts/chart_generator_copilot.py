#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 圖表自動生成器
讀取 raw/ 資料夾的 5 個 Excel，產生 8 個圖表報表 + 1 個整合儀表板到 output/
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    LineChart, BarChart, PieChart, RadarChart, ScatterChart,
    DoughnutChart, Reference, Series,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel, DataLabelList

# ═══════════════════════════════════════════════════════════
# 顏色常數
# ═══════════════════════════════════════════════════════════
C_BLUE   = "2F5496"
C_RED    = "C00000"
C_GREEN  = "548235"
C_GOLD   = "BF8F00"
C_PURPLE = "7030A0"
C_SKY    = "00B0F0"
C_YELLOW = "FFC000"
C_ORANGE = "FF6600"
C_GRAY   = "A5A5A5"
C_LBLUE  = "B4C6E7"
C_WHITE  = "FFFFFF"

PALETTE = [C_BLUE, C_RED, C_GREEN, C_GOLD, C_PURPLE, C_SKY, C_YELLOW, C_ORANGE]

# ═══════════════════════════════════════════════════════════
# 共用格式工具
# ═══════════════════════════════════════════════════════════

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_style(ws, row: int, ncols: int, bg_color: str = C_BLUE):
    """表頭：粗體白字 11pt、指定底色、置中、細框線"""
    fill   = PatternFill(fill_type="solid", fgColor=bg_color)
    font   = Font(bold=True, color=C_WHITE, size=11)
    align  = Alignment(horizontal="center", vertical="center")
    border = _thin_border()
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, align, border


def apply_data_rows(ws, start_row: int, end_row: int, ncols: int,
                    num_cols: list = None):
    """資料列：細框線 + 置中；num_cols 清單中的欄套千分位格式"""
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center")
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border  = border
            cell.alignment = center
            if num_cols and col in num_cols:
                cell.number_format = "#,##0"


def auto_col_width(ws):
    """依內容自動調整欄寬（中文字寬 × 2）"""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_w = 0
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            w = sum(2 if ord(c) > 127 else 1 for c in val)
            if w > max_w:
                max_w = w
        ws.column_dimensions[col_letter].width = min(max_w + 2, 42)


def write_df(ws, df: pd.DataFrame, start_row: int = 1,
             header_color: str = C_BLUE, num_cols: list = None) -> int:
    """
    將 DataFrame 寫入工作表（含表頭格式）。
    回傳最後一筆資料的列號。
    """
    ncols = len(df.columns)

    # 表頭
    for ci, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=col_name)
    apply_header_style(ws, start_row, ncols, header_color)

    # 資料
    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    last_row = start_row + len(df)
    apply_data_rows(ws, start_row + 1, last_row, ncols, num_cols)
    auto_col_width(ws)
    return last_row


# ═══════════════════════════════════════════════════════════
# 路徑設定
# ═══════════════════════════════════════════════════════════
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR  = os.path.join(BASE_DIR, "raw")
OUT_DIR  = os.path.join(BASE_DIR, "output")
os.makedirs(OUT_DIR, exist_ok=True)


def load_data():
    monthly = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
    sales   = pd.read_excel(os.path.join(RAW_DIR, "salesperson_performance.xlsx"))
    market  = pd.read_excel(os.path.join(RAW_DIR, "market_share.xlsx"))
    survey  = pd.read_excel(os.path.join(RAW_DIR, "customer_survey.xlsx"))
    budget  = pd.read_excel(os.path.join(RAW_DIR, "budget_vs_actual.xlsx"))
    return monthly, sales, market, survey, budget


# ═══════════════════════════════════════════════════════════
# 圖表 1：月度營收趨勢折線圖
# ═══════════════════════════════════════════════════════════
def chart1_monthly_trend(monthly: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "月度營收趨勢"

    pivot = (
        monthly.groupby(["年月", "區域"])["營收"]
        .sum().unstack("區域").reset_index()
    )
    pivot.columns.name = None

    ncols  = len(pivot.columns)
    nrows  = len(pivot)
    ncols_data = ncols - 1  # 資料欄（不含年月）
    write_df(ws, pivot, start_row=1, header_color=C_BLUE,
             num_cols=list(range(2, ncols + 1)))

    chart = LineChart()
    chart.title  = "2025 年各區域月度營收趨勢"
    chart.style  = 10
    chart.y_axis.numFmt  = "#,##0"
    chart.y_axis.title   = "營收"
    chart.x_axis.title   = "月份"
    chart.width  = 28
    chart.height = 15

    dates = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)
    for i, region in enumerate(pivot.columns[1:]):
        col  = i + 2
        vals = Reference(ws, min_col=col, min_row=1, max_row=nrows + 1)
        s    = Series(vals, title_from_data=True)
        s.graphicalProperties.line.solidFill = PALETTE[i % len(PALETTE)]
        s.graphicalProperties.line.width     = 25000
        chart.series.append(s)

    chart.set_categories(dates)
    ws.add_chart(chart, "A15")

    path = os.path.join(OUT_DIR, "01_monthly_revenue_trend.xlsx")
    wb.save(path)
    print(f"  ✓ 01_monthly_revenue_trend.xlsx")
    return pivot


# ═══════════════════════════════════════════════════════════
# 圖表 2：區域產品銷售群組柱狀圖
# ═══════════════════════════════════════════════════════════
def chart2_region_product(monthly: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "區域產品銷售"

    pivot = (
        monthly.groupby(["區域", "產品類別"])["營收"]
        .sum().unstack("產品類別").reset_index()
    )
    pivot.columns.name = None

    nrows = len(pivot)
    ncols = len(pivot.columns)
    write_df(ws, pivot, start_row=1, header_color=C_GREEN,
             num_cols=list(range(2, ncols + 1)))

    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "clustered"
    chart.title     = "各區域產品銷售營收比較"
    chart.y_axis.numFmt = "#,##0"
    chart.x_axis.title  = "區域"
    chart.y_axis.title  = "營收"
    chart.width  = 28
    chart.height = 15

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)
    for i, prod in enumerate(pivot.columns[1:]):
        col  = i + 2
        vals = Reference(ws, min_col=col, min_row=1, max_row=nrows + 1)
        s    = Series(vals, title_from_data=True)
        s.graphicalProperties.solidFill = PALETTE[i % len(PALETTE)]
        chart.series.append(s)

    chart.set_categories(cats)
    ws.add_chart(chart, "A8")

    path = os.path.join(OUT_DIR, "02_region_product_sales.xlsx")
    wb.save(path)
    print(f"  ✓ 02_region_product_sales.xlsx")


# ═══════════════════════════════════════════════════════════
# 圖表 3：市占率圓餅圖 + 環圈圖
# ═══════════════════════════════════════════════════════════
def chart3_market_share(market: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "市占率"

    nrows = len(market)
    write_df(ws, market, start_row=1, header_color=C_GOLD)

    pie_colors = [C_BLUE, C_RED, C_GREEN, C_GOLD, C_PURPLE, C_GRAY]
    labels = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)
    data   = Reference(ws, min_col=2, min_row=1, max_row=nrows + 1)

    def _make_labels(series_obj):
        lbl = DataLabelList()
        lbl.showCatName  = True
        lbl.showPercent  = True
        lbl.showVal      = False
        lbl.showSerName  = False
        lbl.showLegendKey= False
        series_obj.dLbls = lbl

    # ── 圓餅圖（左 A10）──────────────────────────────────
    pie = PieChart()
    pie.title  = "市場佔有率"
    pie.width  = 15
    pie.height = 15
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    s_pie = pie.series[0]
    for i, color in enumerate(pie_colors[:nrows]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        s_pie.dPt.append(pt)
    _make_labels(s_pie)
    ws.add_chart(pie, "A10")

    # ── 環圈圖（右 K10）──────────────────────────────────
    donut = DoughnutChart()
    donut.title  = "市場佔有率（環圈）"
    donut.width  = 15
    donut.height = 15
    data2 = Reference(ws, min_col=2, min_row=1, max_row=nrows + 1)
    donut.add_data(data2, titles_from_data=True)
    donut.set_categories(labels)
    s_donut = donut.series[0]
    _make_labels(s_donut)
    ws.add_chart(donut, "K10")

    path = os.path.join(OUT_DIR, "03_market_share_pie.xlsx")
    wb.save(path)
    print(f"  ✓ 03_market_share_pie.xlsx")


# ═══════════════════════════════════════════════════════════
# 圖表 4：業務員績效排名水平長條圖
# ═══════════════════════════════════════════════════════════
def chart4_salesperson_ranking(sales: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "業績排名"

    df = (
        sales[["業務員", "所屬區域", "年度目標", "實際業績", "達成率"]]
        .sort_values("實際業績", ascending=True)
        .reset_index(drop=True)
    )

    nrows = len(df)
    write_df(ws, df, start_row=1, header_color=C_BLUE, num_cols=[3, 4])

    chart = BarChart()
    chart.type           = "bar"       # 水平長條
    chart.title          = "業務員年度業績排名"
    chart.x_axis.numFmt  = "#,##0"
    chart.x_axis.title   = "業績"
    chart.y_axis.title   = "業務員"
    chart.width  = 22
    chart.height = 15

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)
    vals = Reference(ws, min_col=4, min_row=1, max_row=nrows + 1)  # 實際業績
    s    = Series(vals, title_from_data=True)
    s.graphicalProperties.solidFill = C_BLUE
    chart.series.append(s)
    chart.set_categories(cats)

    ws.add_chart(chart, "F1")

    path = os.path.join(OUT_DIR, "04_salesperson_ranking.xlsx")
    wb.save(path)
    print(f"  ✓ 04_salesperson_ranking.xlsx")


# ═══════════════════════════════════════════════════════════
# 圖表 5：預算 vs 實際組合圖（柱狀 + 折線副軸）
# ═══════════════════════════════════════════════════════════
def chart5_budget_vs_actual(budget: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "預算對比"

    nrows = len(budget)
    ncols = len(budget.columns)
    write_df(ws, budget, start_row=1, header_color=C_GREEN, num_cols=[2, 3, 4])

    # ── 主軸柱狀（預算 + 實際）────────────────────────────
    bar = BarChart()
    bar.type     = "col"
    bar.grouping = "clustered"
    bar.title    = "月度預算 vs 實際金額"
    bar.y_axis.numFmt = "#,##0"
    bar.y_axis.title  = "金額"
    bar.x_axis.title  = "月份"
    bar.width    = 28
    bar.height   = 15

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)

    b_vals = Reference(ws, min_col=2, min_row=1, max_row=nrows + 1)
    s_b    = Series(b_vals, title_from_data=True)
    s_b.graphicalProperties.solidFill = C_LBLUE
    bar.series.append(s_b)

    a_vals = Reference(ws, min_col=3, min_row=1, max_row=nrows + 1)
    s_a    = Series(a_vals, title_from_data=True)
    s_a.graphicalProperties.solidFill = C_BLUE
    bar.series.append(s_a)

    bar.set_categories(cats)

    # ── 副軸折線（差異率）─────────────────────────────────
    line = LineChart()
    line.y_axis.axId    = 200
    line.y_axis.crosses = "max"
    line.y_axis.numFmt  = "0.0"
    line.y_axis.title   = "差異率(%)"

    d_vals = Reference(ws, min_col=5, min_row=1, max_row=nrows + 1)
    s_d    = Series(d_vals, title_from_data=True)
    s_d.graphicalProperties.line.solidFill = C_RED
    s_d.graphicalProperties.line.width     = 25000
    line.series.append(s_d)

    bar += line
    ws.add_chart(bar, "A16")

    path = os.path.join(OUT_DIR, "05_budget_vs_actual_combo.xlsx")
    wb.save(path)
    print(f"  ✓ 05_budget_vs_actual_combo.xlsx")


# ═══════════════════════════════════════════════════════════
# 圖表 6：客戶滿意度雷達圖
# ═══════════════════════════════════════════════════════════
def chart6_radar(survey: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "客戶滿意度"

    nrows = len(survey)
    ncols = len(survey.columns)
    write_df(ws, survey, start_row=1, header_color=C_GOLD)

    chart = RadarChart()
    chart.type   = "marker"
    chart.title  = "各區域客戶滿意度雷達圖"
    chart.width  = 20
    chart.height = 15

    # 維度標籤（第 2～最後欄的表頭）
    cats = Reference(ws, min_col=2, max_col=ncols, min_row=1)

    for i, region in enumerate(survey["區域"]):
        row  = i + 2
        vals = Reference(ws, min_col=2, max_col=ncols, min_row=row)
        s    = Series(vals, title=region)
        s.graphicalProperties.line.solidFill = PALETTE[i % len(PALETTE)]
        chart.series.append(s)

    chart.set_categories(cats)
    ws.add_chart(chart, f"A{nrows + 3}")

    path = os.path.join(OUT_DIR, "06_customer_satisfaction_radar.xlsx")
    wb.save(path)
    print(f"  ✓ 06_customer_satisfaction_radar.xlsx")


# ═══════════════════════════════════════════════════════════
# 圖表 7：產品營收 vs 毛利率散佈圖
# ═══════════════════════════════════════════════════════════
def chart7_scatter(monthly: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "產品散佈圖"

    agg = monthly.groupby("產品類別")[["營收", "毛利"]].sum().reset_index()
    agg["毛利率(%)"] = (agg["毛利"] / agg["營收"] * 100).round(1)
    df = agg[["產品類別", "營收", "毛利率(%)"]].copy()

    nrows = len(df)
    write_df(ws, df, start_row=1, header_color=C_BLUE, num_cols=[2])

    chart = ScatterChart()
    chart.title          = "產品營收 vs 毛利率"
    chart.style          = 10
    chart.x_axis.title   = "總營收"
    chart.y_axis.title   = "毛利率(%)"
    chart.x_axis.numFmt  = "#,##0"
    chart.y_axis.numFmt  = "0.0"
    chart.width  = 20
    chart.height = 15

    xvals = Reference(ws, min_col=2, min_row=2, max_row=nrows + 1)
    yvals = Reference(ws, min_col=3, min_row=2, max_row=nrows + 1)

    s = Series(yvals, xvals, title="毛利率")
    s.marker.symbol = "circle"
    s.marker.size   = 8
    s.marker.graphicalProperties.solidFill = C_BLUE
    s.marker.graphicalProperties.line.solidFill = C_BLUE
    s.graphicalProperties.line.noFill = True   # 不連線

    # 顯示 Y 值標籤
    lbl = DataLabelList()
    lbl.showVal      = True
    lbl.showCatName  = False
    lbl.showSerName  = False
    lbl.showLegendKey= False
    s.dLbls = lbl

    chart.series.append(s)
    ws.add_chart(chart, "A9")

    path = os.path.join(OUT_DIR, "07_revenue_profit_scatter.xlsx")
    wb.save(path)
    print(f"  ✓ 07_revenue_profit_scatter.xlsx")


# ═══════════════════════════════════════════════════════════
# 圖表 8：區域季度堆疊柱狀圖
# ═══════════════════════════════════════════════════════════
def chart8_quarterly_stacked(monthly: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "季度堆疊"

    df = monthly.copy()
    df["月"] = df["年月"].astype(str).str[-2:].astype(int)
    df["季度"] = df["月"].apply(lambda m: f"Q{(m - 1) // 3 + 1}")

    pivot = (
        df.groupby(["季度", "區域"])["營收"]
        .sum().unstack("區域").reset_index()
    )
    pivot.columns.name = None

    nrows = len(pivot)
    ncols = len(pivot.columns)
    write_df(ws, pivot, start_row=1, header_color=C_GREEN,
             num_cols=list(range(2, ncols + 1)))

    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "stacked"
    chart.title     = "各區域季度堆疊柱狀圖"
    chart.y_axis.numFmt = "#,##0"
    chart.x_axis.title  = "季度"
    chart.y_axis.title  = "營收"
    chart.width  = 20
    chart.height = 15

    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)
    for i, region in enumerate(pivot.columns[1:]):
        col  = i + 2
        vals = Reference(ws, min_col=col, min_row=1, max_row=nrows + 1)
        s    = Series(vals, title_from_data=True)
        s.graphicalProperties.solidFill = PALETTE[i % len(PALETTE)]
        chart.series.append(s)

    chart.set_categories(cats)
    ws.add_chart(chart, "A8")

    path = os.path.join(OUT_DIR, "08_quarterly_stacked_bar.xlsx")
    wb.save(path)
    print(f"  ✓ 08_quarterly_stacked_bar.xlsx")


# ═══════════════════════════════════════════════════════════
# 整合儀表板 chart_dashboard.xlsx（4 工作表）
# ═══════════════════════════════════════════════════════════
def create_dashboard(monthly, sales, market, budget):
    wb = Workbook()
    wb.remove(wb.active)   # 移除預設空白工作表

    # ── 工作表 1：月度趨勢 ────────────────────────────────
    ws1 = wb.create_sheet("月度趨勢")
    pivot1 = (
        monthly.groupby(["年月", "區域"])["營收"]
        .sum().unstack("區域").reset_index()
    )
    pivot1.columns.name = None
    nrows1 = len(pivot1)
    write_df(ws1, pivot1, start_row=1, header_color=C_BLUE,
             num_cols=list(range(2, len(pivot1.columns) + 1)))

    c1 = LineChart()
    c1.title  = "2025 年各區域月度營收趨勢"
    c1.y_axis.numFmt = "#,##0"
    c1.width  = 28
    c1.height = 15
    dates1 = Reference(ws1, min_col=1, min_row=2, max_row=nrows1 + 1)
    for i, region in enumerate(pivot1.columns[1:]):
        v = Reference(ws1, min_col=i + 2, min_row=1, max_row=nrows1 + 1)
        s = Series(v, title_from_data=True)
        s.graphicalProperties.line.solidFill = PALETTE[i % len(PALETTE)]
        s.graphicalProperties.line.width     = 25000
        c1.series.append(s)
    c1.set_categories(dates1)
    ws1.add_chart(c1, "A15")

    # ── 工作表 2：市占率 ──────────────────────────────────
    ws2 = wb.create_sheet("市占率")
    nrows2 = len(market)
    write_df(ws2, market, start_row=1, header_color=C_GOLD)
    labels2 = Reference(ws2, min_col=1, min_row=2, max_row=nrows2 + 1)
    data2   = Reference(ws2, min_col=2, min_row=1, max_row=nrows2 + 1)

    c2 = PieChart()
    c2.title  = "市場佔有率"
    c2.width  = 15
    c2.height = 15
    c2.add_data(data2, titles_from_data=True)
    c2.set_categories(labels2)
    pie_colors = [C_BLUE, C_RED, C_GREEN, C_GOLD, C_PURPLE, C_GRAY]
    s2 = c2.series[0]
    for i, color in enumerate(pie_colors[:nrows2]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        s2.dPt.append(pt)
    lbl2 = DataLabelList()
    lbl2.showCatName = True
    lbl2.showPercent = True
    lbl2.showVal     = False
    s2.dLbls = lbl2
    ws2.add_chart(c2, "A8")

    # ── 工作表 3：預算對比 ────────────────────────────────
    ws3 = wb.create_sheet("預算對比")
    nrows3 = len(budget)
    write_df(ws3, budget, start_row=1, header_color=C_GREEN, num_cols=[2, 3, 4])

    bar3 = BarChart()
    bar3.type = "col"; bar3.grouping = "clustered"
    bar3.title = "月度預算 vs 實際金額"
    bar3.y_axis.numFmt = "#,##0"
    bar3.width = 28; bar3.height = 15
    cats3 = Reference(ws3, min_col=1, min_row=2, max_row=nrows3 + 1)
    sb  = Series(Reference(ws3, min_col=2, min_row=1, max_row=nrows3 + 1), title_from_data=True)
    sb.graphicalProperties.solidFill = C_LBLUE
    bar3.series.append(sb)
    sa  = Series(Reference(ws3, min_col=3, min_row=1, max_row=nrows3 + 1), title_from_data=True)
    sa.graphicalProperties.solidFill = C_BLUE
    bar3.series.append(sa)
    bar3.set_categories(cats3)
    line3 = LineChart()
    line3.y_axis.axId    = 200
    line3.y_axis.crosses = "max"
    line3.y_axis.numFmt  = "0.0"
    line3.y_axis.title   = "差異率(%)"
    sd  = Series(Reference(ws3, min_col=5, min_row=1, max_row=nrows3 + 1), title_from_data=True)
    sd.graphicalProperties.line.solidFill = C_RED
    sd.graphicalProperties.line.width     = 25000
    line3.series.append(sd)
    bar3 += line3
    ws3.add_chart(bar3, "A16")

    # ── 工作表 4：績效排名 ────────────────────────────────
    ws4 = wb.create_sheet("績效排名")
    df4 = (
        sales[["業務員", "所屬區域", "實際業績", "達成率"]]
        .sort_values("實際業績", ascending=True)
        .reset_index(drop=True)
    )
    nrows4 = len(df4)
    write_df(ws4, df4, start_row=1, header_color=C_BLUE, num_cols=[3])
    c4 = BarChart()
    c4.type  = "bar"
    c4.title = "業務員年度業績排名"
    c4.width = 20; c4.height = 15
    c4_cats = Reference(ws4, min_col=1, min_row=2, max_row=nrows4 + 1)
    c4_vals = Reference(ws4, min_col=3, min_row=1, max_row=nrows4 + 1)
    s4 = Series(c4_vals, title_from_data=True)
    s4.graphicalProperties.solidFill = C_BLUE
    c4.series.append(s4)
    c4.set_categories(c4_cats)
    ws4.add_chart(c4, "F1")

    path = os.path.join(OUT_DIR, "chart_dashboard.xlsx")
    wb.save(path)
    print(f"  ✓ chart_dashboard.xlsx")


# ═══════════════════════════════════════════════════════════
# 主程式
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("📊 Excel 圖表產生器")
    print("=" * 55)
    print("📂 讀取 raw/ 資料...")
    monthly, sales, market, survey, budget = load_data()

    print("\n🔧 產生圖表報表...")
    chart1_monthly_trend(monthly)
    chart2_region_product(monthly)
    chart3_market_share(market)
    chart4_salesperson_ranking(sales)
    chart5_budget_vs_actual(budget)
    chart6_radar(survey)
    chart7_scatter(monthly)
    chart8_quarterly_stacked(monthly)

    print("\n📊 產生整合儀表板...")
    create_dashboard(monthly, sales, market, budget)

    print("\n" + "=" * 55)
    print("✅ 完成！output/ 產出檔案：")
    outputs = [
        ("01_monthly_revenue_trend.xlsx",      "月度營收趨勢折線圖"),
        ("02_region_product_sales.xlsx",        "區域產品銷售群組柱狀圖"),
        ("03_market_share_pie.xlsx",            "市占率圓餅圖 + 環圈圖"),
        ("04_salesperson_ranking.xlsx",         "業務員績效排名水平長條圖"),
        ("05_budget_vs_actual_combo.xlsx",      "預算 vs 實際組合圖"),
        ("06_customer_satisfaction_radar.xlsx", "客戶滿意度雷達圖"),
        ("07_revenue_profit_scatter.xlsx",      "產品營收 vs 毛利率散佈圖"),
        ("08_quarterly_stacked_bar.xlsx",       "區域季度堆疊柱狀圖"),
        ("chart_dashboard.xlsx",                "整合儀表板（4 工作表）"),
    ]
    for fname, desc in outputs:
        print(f"  📄 {fname:<42} {desc}")
