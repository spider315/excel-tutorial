"""
demo/single_chart_example.py
最簡單的 openpyxl 圖表範例 — 教學用
═══════════════════════════════════════
用途：讓學員理解 openpyxl 建立圖表的基本流程
執行：python demo/single_chart_example.py
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── 步驟 1：建立工作簿與工作表 ─────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "簡單範例"

# ── 步驟 2：寫入資料 ─────────────────────────────────────
data = [
    ["水果", "銷售量"],
    ["蘋果", 150],
    ["香蕉", 200],
    ["橘子", 120],
    ["葡萄", 80],
    ["西瓜", 60],
]

for row in data:
    ws.append(row)

# ── 步驟 3：套用表頭樣式 ─────────────────────────────────
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ── 步驟 4：建立柱狀圖 ─────────────────────────────────
chart = BarChart()
chart.type = "col"                       # 垂直柱狀
chart.grouping = "clustered"             # 群組模式
chart.title = "水果銷售量比較"
chart.x_axis.title = "水果"
chart.y_axis.title = "銷售量"
chart.width = 20                         # 圖表寬度
chart.height = 12                        # 圖表高度
chart.style = 10                         # 內建樣式編號

# ── 步驟 5：指定資料範圍 ─────────────────────────────────
# data_ref: 數值資料（B1:B6，包含標題）
data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=6)
# cats_ref: 類別標籤（A2:A6，不含標題）
cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

# ── 步驟 6：自訂顏色 ─────────────────────────────────────
chart.series[0].graphicalProperties.solidFill = "2F5496"

# ── 步驟 7：放置圖表到工作表 ──────────────────────────────
ws.add_chart(chart, "D2")                # 放在 D2 儲存格

# ── 步驟 8：儲存檔案 ─────────────────────────────────────
output_path = os.path.join(OUTPUT_DIR, "demo_simple_chart.xlsx")
wb.save(output_path)
print(f"✅ 範例圖表已儲存：{output_path}")
print()
print("📝 學習重點：")
print("   1. Workbook() → 建立工作簿")
print("   2. ws.append() → 寫入資料")
print("   3. BarChart() → 建立圖表物件")
print("   4. Reference() → 指定資料範圍")
print("   5. add_data() → 綁定資料到圖表")
print("   6. set_categories() → 設定 X 軸標籤")
print("   7. add_chart() → 放置圖表到工作表")
