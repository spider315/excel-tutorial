import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import (
    LineChart, BarChart, PieChart, DoughnutChart,
    RadarChart, ScatterChart, Series, Reference
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import Marker

# 設定路徑
RAW_DIR = "raw"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 顏色定義
COLORS = {
    'dark_blue': '2F5496',
    'dark_red': 'C00000',
    'dark_green': '548235',
    'dark_gold': 'BF8F00',
    'purple': '7030A0',
    'sky_blue': '00B0F0',
    'golden_yellow': 'FFC000',
    'orange': 'FF6600',
    'light_blue': 'B4C6E7',
    'white': 'FFFFFF',
    'grey': 'A5A5A5'
}

def apply_common_style(ws, header_color_hex='2F5496'):
    """套用共用格式：表頭、框線、自動欄寬"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color=header_color_hex, end_color=header_color_hex, fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 找到資料範圍
    max_row = ws.max_row
    max_col = ws.max_column
    
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
    
    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def create_report_01():
    """圖表 1：月度營收趨勢折線圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
    pivot_df = df.pivot_table(index='年月', columns='區域', values='營收', aggfunc='sum')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Revenue Trend"
    
    for r in dataframe_to_rows(pivot_df.reset_index(), index=False, header=True):
        ws.append(r)
    
    apply_common_style(ws, COLORS['dark_blue'])
    
    chart = LineChart()
    chart.title = "2025 年各區域月度營收趨勢"
    chart.style = 13
    chart.y_axis.title = "營收"
    chart.x_axis.title = "月份"
    chart.y_axis.number_format = '#,##0'
    
    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # 設定折線寬度 (25000 EMU)
    for s in chart.series:
        s.graphicalProperties.line.width = 25000
        
    chart.width = 28
    chart.height = 15
    ws.add_chart(chart, "A15")
    
    wb.save(os.path.join(OUTPUT_DIR, "01_monthly_revenue_trend.xlsx"))
    return wb

def create_report_02():
    """圖表 2：區域產品銷售柱狀圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
    pivot_df = df.pivot_table(index='區域', columns='產品類別', values='營收', aggfunc='sum')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Region Product Sales"
    
    for r in dataframe_to_rows(pivot_df.reset_index(), index=False, header=True):
        ws.append(r)
    
    apply_common_style(ws, COLORS['dark_green']) # 表頭用綠色底
    
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "各區域產品銷售營收比較"
    chart.y_axis.title = "營收"
    chart.x_axis.title = "區域"
    
    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A8")
    wb.save(os.path.join(OUTPUT_DIR, "02_region_product_sales.xlsx"))
    return wb

def create_report_03():
    """圖表 3：市占率圓餅圖與環圈圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "market_share.xlsx"))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Share"
    
    # 只要品牌和市占率
    data_df = df[['品牌', '市占率(%)']]
    for r in dataframe_to_rows(data_df, index=False, header=True):
        ws.append(r)
        
    apply_common_style(ws, COLORS['dark_gold']) # 表頭用橘色底
    
    # 圓餅圖
    pie = PieChart()
    pie.title = "品牌市占率圓餅圖"
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    pie.add_data(data)
    pie.set_categories(cats)
    pie.dLbls = DataLabelList()
    pie.dLbls.showCatName = True
    pie.dLbls.showPercent = True
    
    # 自訂顏色
    slice_colors = [COLORS['dark_blue'], COLORS['dark_red'], COLORS['dark_green'], 
                    COLORS['dark_gold'], COLORS['purple'], COLORS['grey']]
    for i, color in enumerate(slice_colors):
        if i < len(data_df):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)
            
    ws.add_chart(pie, "A10")
    
    # 環圈圖
    doughnut = DoughnutChart()
    doughnut.title = "品牌市占率環圈圖"
    doughnut.add_data(data)
    doughnut.set_categories(cats)
    doughnut.dLbls = DataLabelList()
    doughnut.dLbls.showCatName = True
    doughnut.dLbls.showPercent = True
    
    ws.add_chart(doughnut, "K10")
    
    wb.save(os.path.join(OUTPUT_DIR, "03_market_share_pie.xlsx"))
    return wb

def create_report_04():
    """圖表 4：業務員績效排名長條圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "salesperson_performance.xlsx"))
    df_sorted = df[['業務員', '實際業績']].sort_values(by='實際業績', ascending=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Salesperson Ranking"
    
    for r in dataframe_to_rows(df_sorted, index=False, header=True):
        ws.append(r)
        
    apply_common_style(ws, COLORS['dark_blue'])
    
    chart = BarChart()
    chart.type = "bar" # 水平長條圖
    chart.title = "業務員年度業績排名"
    chart.x_axis.title = "業務員"
    chart.y_axis.title = "業績"
    
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # 長條用深藍色
    s = chart.series[0]
    s.graphicalProperties.solidFill = COLORS['dark_blue']
    
    ws.add_chart(chart, "F1")
    wb.save(os.path.join(OUTPUT_DIR, "04_salesperson_ranking.xlsx"))
    return wb

def create_report_05():
    """圖表 5：預算 vs 實際組合圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "budget_vs_actual.xlsx"))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget vs Actual"
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    apply_common_style(ws, COLORS['dark_green'])
    
    # 柱狀圖 (預算 vs 實際)
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "月度預算 vs 實際金額"
    bar_chart.y_axis.title = "金額"
    
    # 預算金額 (B) & 實際金額 (C)
    data_bar = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    bar_chart.add_data(data_bar, titles_from_data=True)
    bar_chart.set_categories(cats)
    
    # 設定顏色
    bar_chart.series[0].graphicalProperties.solidFill = COLORS['light_blue'] # 預算
    bar_chart.series[1].graphicalProperties.solidFill = COLORS['dark_blue']  # 實際
    
    # 折線圖 (差異率 E)
    line_chart = LineChart()
    data_line = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
    line_chart.add_data(data_line, titles_from_data=True)
    line_chart.y_axis.axId = 200 # 副軸
    line_chart.y_axis.title = "差異率 (%)"
    line_chart.y_axis.crosses = "max" # 副軸放在右邊
    
    # 暗紅色線條
    s_line = line_chart.series[0]
    s_line.graphicalProperties.line.solidFill = COLORS['dark_red']
    
    # 組合
    bar_chart += line_chart
    
    ws.add_chart(bar_chart, "A16")
    wb.save(os.path.join(OUTPUT_DIR, "05_budget_vs_actual_combo.xlsx"))
    return wb

def create_report_06():
    """圖表 6：客戶滿意度雷達圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "customer_survey.xlsx"))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Satisfaction"
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    apply_common_style(ws, COLORS['dark_gold'])
    
    chart = RadarChart()
    chart.type = "marker"
    chart.title = "各區域客戶滿意度雷達圖"
    
    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A8")
    wb.save(os.path.join(OUTPUT_DIR, "06_customer_satisfaction_radar.xlsx"))
    return wb

def create_report_07():
    """圖表 7：產品營收 vs 毛利率散佈圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
    summary = df.groupby('產品類別').agg({'營收': 'sum', '毛利': 'sum'})
    summary['毛利率(%)'] = (summary['毛利'] / summary['營收'] * 100).round(2)
    summary = summary.reset_index()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue vs Profit Margin"
    
    # 寫入資料：產品類別, 營收, 毛利率
    ws.append(['產品類別', '總營收', '毛利率(%)'])
    for idx, row in summary.iterrows():
        ws.append([row['產品類別'], row['營收'], row['毛利率(%)']])
        
    apply_common_style(ws, COLORS['dark_blue'])
    
    chart = ScatterChart()
    chart.title = "產品營收 vs 毛利率散佈圖"
    chart.x_axis.title = "總營收"
    chart.y_axis.title = "毛利率(%)"
    chart.x_axis.number_format = '#,##0'
    
    xvalues = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    yvalues = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    series = Series(yvalues, xvalues, title_from_data=False)
    
    # 散佈點不連線 (預設即不連線，除非設定為 smooth 或 straight)
    chart.series.append(series)
    
    # 顯示 Y 值資料標籤
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    
    ws.add_chart(chart, "A9")
    wb.save(os.path.join(OUTPUT_DIR, "07_revenue_profit_scatter.xlsx"))
    return wb

def create_report_08():
    """圖表 8：區域季度堆疊柱狀圖"""
    df = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
    
    # 算出季度
    df['月份'] = pd.to_datetime(df['年月']).dt.month
    df['季度'] = df['月份'].apply(lambda x: f"Q{(x-1)//3 + 1}")
    
    pivot_df = df.pivot_table(index='季度', columns='區域', values='營收', aggfunc='sum')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarterly Stacked Bar"
    
    for r in dataframe_to_rows(pivot_df.reset_index(), index=False, header=True):
        ws.append(r)
        
    apply_common_style(ws, COLORS['dark_green'])
    
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "各季度各區域營收堆疊圖"
    
    data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A8")
    wb.save(os.path.join(OUTPUT_DIR, "08_quarterly_stacked_bar.xlsx"))
    return wb

def create_dashboard():
    """整合儀表板：chart_dashboard.xlsx"""
    wb = Workbook()
    
    # 1. 月度趨勢
    ws1 = wb.active
    ws1.title = "月度趨勢"
    df1 = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
    pivot1 = df1.pivot_table(index='年月', columns='區域', values='營收', aggfunc='sum')
    for r in dataframe_to_rows(pivot1.reset_index(), index=False, header=True):
        ws1.append(r)
    apply_common_style(ws1, COLORS['dark_blue'])
    
    chart1 = LineChart()
    chart1.title = "2025 年各區域月度營收趨勢"
    data1 = Reference(ws1, min_col=2, min_row=1, max_col=ws1.max_column, max_row=ws1.max_row)
    cats1 = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    for s in chart1.series: s.graphicalProperties.line.width = 25000
    chart1.width, chart1.height = 28, 15
    ws1.add_chart(chart1, "A15")
    
    # 2. 市占率
    ws2 = wb.create_sheet("市占率")
    df2 = pd.read_excel(os.path.join(RAW_DIR, "market_share.xlsx"))[['品牌', '市占率(%)']]
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)
    apply_common_style(ws2, COLORS['dark_gold'])
    
    pie2 = PieChart()
    pie2.title = "品牌市占率"
    data2 = Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row)
    cats2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    pie2.add_data(data2)
    pie2.set_categories(cats2)
    pie2.dLbls = DataLabelList(showCatName=True, showPercent=True)
    ws2.add_chart(pie2, "A10")
    
    # 3. 預算對比
    ws3 = wb.create_sheet("預算對比")
    df3 = pd.read_excel(os.path.join(RAW_DIR, "budget_vs_actual.xlsx"))
    for r in dataframe_to_rows(df3, index=False, header=True):
        ws3.append(r)
    apply_common_style(ws3, COLORS['dark_green'])
    
    bar3 = BarChart()
    bar3.type = "col"
    bar3.title = "月度預算 vs 實際金額"
    data3 = Reference(ws3, min_col=2, min_row=1, max_col=3, max_row=ws3.max_row)
    cats3 = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    bar3.add_data(data3, titles_from_data=True)
    bar3.set_categories(cats3)
    bar3.series[0].graphicalProperties.solidFill = COLORS['light_blue']
    bar3.series[1].graphicalProperties.solidFill = COLORS['dark_blue']
    ws3.add_chart(bar3, "A16")
    
    # 4. 績效排名
    ws4 = wb.create_sheet("績效排名")
    df4 = pd.read_excel(os.path.join(RAW_DIR, "salesperson_performance.xlsx"))[['業務員', '實際業績']].sort_values(by='實際業績', ascending=True)
    for r in dataframe_to_rows(df4, index=False, header=True):
        ws4.append(r)
    apply_common_style(ws4, COLORS['dark_blue'])
    
    bar4 = BarChart()
    bar4.type = "bar"
    bar4.title = "業務員年度業績排名"
    data4 = Reference(ws4, min_col=2, min_row=1, max_row=ws4.max_row)
    cats4 = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
    bar4.add_data(data4, titles_from_data=True)
    bar4.set_categories(cats4)
    bar4.series[0].graphicalProperties.solidFill = COLORS['dark_blue']
    ws4.add_chart(bar4, "F1")
    
    wb.save(os.path.join(OUTPUT_DIR, "chart_dashboard.xlsx"))

if __name__ == "__main__":
    print("🚀 開始產生報表...")
    create_report_01()
    print("✅ 01_monthly_revenue_trend.xlsx 完成")
    create_report_02()
    print("✅ 02_region_product_sales.xlsx 完成")
    create_report_03()
    print("✅ 03_market_share_pie.xlsx 完成")
    create_report_04()
    print("✅ 04_salesperson_ranking.xlsx 完成")
    create_report_05()
    print("✅ 05_budget_vs_actual_combo.xlsx 完成")
    create_report_06()
    print("✅ 06_customer_satisfaction_radar.xlsx 完成")
    create_report_07()
    print("✅ 07_revenue_profit_scatter.xlsx 完成")
    create_report_08()
    print("✅ 08_quarterly_stacked_bar.xlsx 完成")
    create_dashboard()
    print("✅ chart_dashboard.xlsx 完成")
    
    print("\n" + "="*40)
    print("摘要：已成功產出 8 個報表與 1 個儀表板至 output/ 資料夾")
    print("="*40)
