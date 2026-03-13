"""
04-excel-charts / chart_generator.py
讀取 raw/ 資料，自動產生多種專業 Excel 圖表
═══════════════════════════════════════════════
執行方式：python chart_generator.py
輸出位置：output/ 資料夾

圖表清單：
  1. 月度營收趨勢折線圖
  2. 區域產品銷售柱狀圖
  3. 市占率圓餅圖與環圈圖
  4. 業務員績效排名長條圖
  5. 預算 vs 實際組合圖（柱狀＋折線）
  6. 客戶滿意度雷達圖
  7. 產品營收 vs 毛利散佈圖
  8. 區域季度堆疊柱狀圖
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart, BarChart, PieChart, Reference,
    ScatterChart, RadarChart, Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import deepcopy

RAW_DIR = os.path.join(os.path.dirname(__file__), "raw")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── 共用樣式 ─────────────────────────────────────────────────
HEADER_FONT = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=11)
HEADER_FILL_BLUE = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CHART_COLORS = [
    "2F5496", "C00000", "548235", "BF8F00",
    "7030A0", "00B0F0", "FFC000", "FF6600",
]


def style_header(ws, fill=HEADER_FILL_BLUE):
    """套用表頭樣式"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cells(ws):
    """套用資料區域框線"""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_column_width(ws):
    """自動調整欄寬"""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


# ══════════════════════════════════════════════════════════════
# 讀取原始資料
# ══════════════════════════════════════════════════════════════

print("📂 讀取原始資料...")
df_sales = pd.read_excel(os.path.join(RAW_DIR, "monthly_sales_detail.xlsx"))
df_perf = pd.read_excel(os.path.join(RAW_DIR, "salesperson_performance.xlsx"))
df_share = pd.read_excel(os.path.join(RAW_DIR, "market_share.xlsx"))
df_survey = pd.read_excel(os.path.join(RAW_DIR, "customer_survey.xlsx"))
df_budget = pd.read_excel(os.path.join(RAW_DIR, "budget_vs_actual.xlsx"))
print("✅ 資料載入完成\n")


# ══════════════════════════════════════════════════════════════
# 報表 1：月度營收趨勢折線圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 1：月度營收趨勢折線圖...")

monthly_by_region = df_sales.groupby(["年月", "區域"])["營收"].sum().reset_index()
pivot_monthly = monthly_by_region.pivot(index="年月", columns="區域", values="營收")
pivot_monthly = pivot_monthly.reset_index()

wb1 = Workbook()
ws1 = wb1.active
ws1.title = "月度營收趨勢"

# 寫入資料
headers = ["月份"] + list(pivot_monthly.columns[1:])
ws1.append(headers)
for _, row in pivot_monthly.iterrows():
    ws1.append([row["年月"]] + [int(row[r]) for r in pivot_monthly.columns[1:]])

style_header(ws1)
style_data_cells(ws1)
auto_column_width(ws1)

# 建立折線圖
chart1 = LineChart()
chart1.title = "2025 年各區域月度營收趨勢"
chart1.x_axis.title = "月份"
chart1.y_axis.title = "營收（元）"
chart1.y_axis.numFmt = '#,##0'
chart1.width = 28
chart1.height = 15
chart1.style = 10

data_ref = Reference(ws1, min_col=2, max_col=ws1.max_column,
                     min_row=1, max_row=ws1.max_row)
cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)

for i, series in enumerate(chart1.series):
    series.graphicalProperties.line.width = 25000
    if i < len(CHART_COLORS):
        series.graphicalProperties.line.solidFill = CHART_COLORS[i]

ws1.add_chart(chart1, "A15")

wb1.save(os.path.join(OUTPUT_DIR, "01_monthly_revenue_trend.xlsx"))
print("  ✅ 01_monthly_revenue_trend.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 2：區域產品銷售柱狀圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 2：區域產品銷售柱狀圖...")

region_product = df_sales.groupby(["區域", "產品類別"])["營收"].sum().reset_index()
pivot_rp = region_product.pivot(index="區域", columns="產品類別", values="營收")
pivot_rp = pivot_rp.reset_index()

wb2 = Workbook()
ws2 = wb2.active
ws2.title = "區域產品銷售"

headers2 = ["區域"] + list(pivot_rp.columns[1:])
ws2.append(headers2)
for _, row in pivot_rp.iterrows():
    ws2.append([row["區域"]] + [int(row[c]) for c in pivot_rp.columns[1:]])

style_header(ws2, HEADER_FILL_GREEN)
style_data_cells(ws2)
auto_column_width(ws2)

chart2 = BarChart()
chart2.type = "col"
chart2.grouping = "clustered"
chart2.title = "各區域產品銷售營收比較"
chart2.x_axis.title = "區域"
chart2.y_axis.title = "營收（元）"
chart2.y_axis.numFmt = '#,##0'
chart2.width = 26
chart2.height = 15
chart2.style = 10

data_ref2 = Reference(ws2, min_col=2, max_col=ws2.max_column,
                      min_row=1, max_row=ws2.max_row)
cats_ref2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)

for i, series in enumerate(chart2.series):
    if i < len(CHART_COLORS):
        series.graphicalProperties.solidFill = CHART_COLORS[i]

ws2.add_chart(chart2, "A8")

wb2.save(os.path.join(OUTPUT_DIR, "02_region_product_sales.xlsx"))
print("  ✅ 02_region_product_sales.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 3：市占率圓餅圖與環圈圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 3：市占率圓餅圖...")

wb3 = Workbook()
ws3 = wb3.active
ws3.title = "市占率分析"

ws3.append(["品牌", "市占率(%)", "營收(萬元)", "年成長率(%)"])
for _, row in df_share.iterrows():
    ws3.append([row["品牌"], row["市占率(%)"], row["營收(萬元)"], row["年成長率(%)"]])

style_header(ws3, HEADER_FILL_ORANGE)
style_data_cells(ws3)
auto_column_width(ws3)

# 圓餅圖
pie1 = PieChart()
pie1.title = "市場份額分佈"
pie1.width = 18
pie1.height = 14
pie1.style = 10

pie_data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
pie_cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
pie1.add_data(pie_data, titles_from_data=True)
pie1.set_categories(pie_cats)

pie1.dataLabels = DataLabelList()
pie1.dataLabels.showPercent = True
pie1.dataLabels.showCatName = True
pie1.dataLabels.showVal = False

# 自訂每個扇區顏色
pie_colors = ["2F5496", "C00000", "548235", "BF8F00", "7030A0", "A5A5A5"]
for i, color in enumerate(pie_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie1.series[0].data_points.append(pt)

ws3.add_chart(pie1, "A10")

# 環圈圖 (用第二個圓餅圖模擬 — openpyxl 支援 DoughnutChart)
from openpyxl.chart import DoughnutChart
donut = DoughnutChart()
donut.title = "市場份額（環圈圖）"
donut.width = 18
donut.height = 14
donut.style = 10

donut_data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
donut_cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
donut.add_data(donut_data, titles_from_data=True)
donut.set_categories(donut_cats)

donut.dataLabels = DataLabelList()
donut.dataLabels.showPercent = True
donut.dataLabels.showCatName = True

ws3.add_chart(donut, "K10")

wb3.save(os.path.join(OUTPUT_DIR, "03_market_share_pie.xlsx"))
print("  ✅ 03_market_share_pie.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 4：業務員績效排名長條圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 4：業務員績效排名長條圖...")

df_perf_sorted = df_perf.sort_values("實際業績", ascending=True)

wb4 = Workbook()
ws4 = wb4.active
ws4.title = "績效排名"

ws4.append(["業務員", "實際業績", "達成率(%)", "客戶滿意度"])
for _, row in df_perf_sorted.iterrows():
    ws4.append([row["業務員"], int(row["實際業績"]), row["達成率"], row["客戶滿意度"]])

style_header(ws4)
style_data_cells(ws4)
auto_column_width(ws4)

chart4 = BarChart()
chart4.type = "bar"      # 水平長條
chart4.grouping = "clustered"
chart4.title = "業務員年度業績排名"
chart4.x_axis.title = "業務員"
chart4.y_axis.title = "實際業績（元）"
chart4.y_axis.numFmt = '#,##0'
chart4.width = 26
chart4.height = 16
chart4.style = 10

data_ref4 = Reference(ws4, min_col=2, max_col=2, min_row=1, max_row=ws4.max_row)
cats_ref4 = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)

chart4.series[0].graphicalProperties.solidFill = "2F5496"

ws4.add_chart(chart4, "F1")

wb4.save(os.path.join(OUTPUT_DIR, "04_salesperson_ranking.xlsx"))
print("  ✅ 04_salesperson_ranking.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 5：預算 vs 實際組合圖（柱狀＋折線）
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 5：預算 vs 實際組合圖...")

wb5 = Workbook()
ws5 = wb5.active
ws5.title = "預算vs實際"

ws5.append(["月份", "預算金額", "實際金額", "差異率(%)"])
for _, row in df_budget.iterrows():
    ws5.append([row["月份"], int(row["預算金額"]), int(row["實際金額"]), row["差異率(%)"]])

style_header(ws5, HEADER_FILL_GREEN)
style_data_cells(ws5)
auto_column_width(ws5)

# 柱狀圖（預算 & 實際）
bar5 = BarChart()
bar5.type = "col"
bar5.grouping = "clustered"
bar5.title = "月度預算 vs 實際金額"
bar5.y_axis.title = "金額（元）"
bar5.y_axis.numFmt = '#,##0'
bar5.width = 28
bar5.height = 15
bar5.style = 10

bar_data = Reference(ws5, min_col=2, max_col=3, min_row=1, max_row=ws5.max_row)
bar_cats = Reference(ws5, min_col=1, min_row=2, max_row=ws5.max_row)
bar5.add_data(bar_data, titles_from_data=True)
bar5.set_categories(bar_cats)

bar5.series[0].graphicalProperties.solidFill = "B4C6E7"  # 淺藍 = 預算
bar5.series[1].graphicalProperties.solidFill = "2F5496"   # 深藍 = 實際

# 折線圖（差異率）疊加到第二軸
line5 = LineChart()
line5.y_axis.title = "差異率（%）"
line5.y_axis.numFmt = '0.0"%"'
line_data = Reference(ws5, min_col=4, max_col=4, min_row=1, max_row=ws5.max_row)
line5.add_data(line_data, titles_from_data=True)
line5.series[0].graphicalProperties.line.solidFill = "C00000"
line5.series[0].graphicalProperties.line.width = 25000
line5.y_axis.axId = 200

bar5.y_axis.crosses = "min"
bar5 += line5

ws5.add_chart(bar5, "A16")

wb5.save(os.path.join(OUTPUT_DIR, "05_budget_vs_actual_combo.xlsx"))
print("  ✅ 05_budget_vs_actual_combo.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 6：客戶滿意度雷達圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 6：客戶滿意度雷達圖...")

wb6 = Workbook()
ws6 = wb6.active
ws6.title = "滿意度雷達圖"

headers6 = ["區域"] + list(df_survey.columns[1:])
ws6.append(headers6)
for _, row in df_survey.iterrows():
    ws6.append([row["區域"]] + [row[c] for c in df_survey.columns[1:]])

style_header(ws6, HEADER_FILL_ORANGE)
style_data_cells(ws6)
auto_column_width(ws6)

radar = RadarChart()
radar.type = "marker"
radar.title = "各區域客戶滿意度雷達圖"
radar.width = 22
radar.height = 16
radar.style = 10

radar_data = Reference(ws6, min_col=2, max_col=ws6.max_column,
                       min_row=1, max_row=ws6.max_row)
radar_cats = Reference(ws6, min_col=1, min_row=2, max_row=ws6.max_row)

# 雷達圖的資料方向：每列是一個區域，每欄是一個維度
# 需要轉置 — 改用逐列加入
for i in range(2, ws6.max_row + 1):
    values = Reference(ws6, min_col=2, max_col=ws6.max_column, min_row=i, max_row=i)
    series = Series(values, title=ws6.cell(row=i, column=1).value)
    radar.series.append(series)

cats = Reference(ws6, min_col=2, max_col=ws6.max_column, min_row=1, max_row=1)
radar.set_categories(cats)

for i, s in enumerate(radar.series):
    if i < len(CHART_COLORS):
        s.graphicalProperties.line.solidFill = CHART_COLORS[i]
        s.graphicalProperties.line.width = 25000

ws6.add_chart(radar, "A8")

wb6.save(os.path.join(OUTPUT_DIR, "06_customer_satisfaction_radar.xlsx"))
print("  ✅ 06_customer_satisfaction_radar.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 7：產品營收 vs 毛利率散佈圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 7：產品營收 vs 毛利率散佈圖...")

product_summary = df_sales.groupby("產品類別").agg(
    營收=("營收", "sum"),
    毛利=("毛利", "sum"),
    數量=("銷售數量", "sum"),
).reset_index()
product_summary["毛利率(%)"] = round(product_summary["毛利"] / product_summary["營收"] * 100, 1)

wb7 = Workbook()
ws7 = wb7.active
ws7.title = "營收毛利散佈圖"

ws7.append(["產品類別", "總營收", "總毛利", "毛利率(%)", "總銷售數量"])
for _, row in product_summary.iterrows():
    ws7.append([row["產品類別"], int(row["營收"]), int(row["毛利"]),
                row["毛利率(%)"], int(row["數量"])])

style_header(ws7)
style_data_cells(ws7)
auto_column_width(ws7)

scatter = ScatterChart()
scatter.title = "產品營收 vs 毛利率分析"
scatter.x_axis.title = "總營收（元）"
scatter.x_axis.numFmt = '#,##0'
scatter.y_axis.title = "毛利率（%）"
scatter.y_axis.numFmt = '0.0"%"'
scatter.width = 24
scatter.height = 15
scatter.style = 10

x_values = Reference(ws7, min_col=2, min_row=2, max_row=ws7.max_row)
y_values = Reference(ws7, min_col=4, min_row=2, max_row=ws7.max_row)
scatter_series = Series(y_values, x_values, title="產品")
scatter_series.graphicalProperties.line.noFill = True
scatter.series.append(scatter_series)

# 加入資料標籤
scatter_series.dLbls = DataLabelList()
scatter_series.dLbls.showSerName = False
scatter_series.dLbls.showVal = True

ws7.add_chart(scatter, "A9")

wb7.save(os.path.join(OUTPUT_DIR, "07_revenue_profit_scatter.xlsx"))
print("  ✅ 07_revenue_profit_scatter.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 報表 8：區域季度堆疊柱狀圖
# ══════════════════════════════════════════════════════════════
print("📊 產生報表 8：區域季度堆疊柱狀圖...")

df_sales["季度"] = pd.to_datetime(df_sales["年月"]).dt.quarter.map(
    {1: "Q1", 2: "Q2", 3: "Q3", 4: "Q4"}
)
quarterly = df_sales.groupby(["季度", "區域"])["營收"].sum().reset_index()
pivot_q = quarterly.pivot(index="季度", columns="區域", values="營收").reset_index()

wb8 = Workbook()
ws8 = wb8.active
ws8.title = "季度堆疊圖"

headers8 = ["季度"] + list(pivot_q.columns[1:])
ws8.append(headers8)
for _, row in pivot_q.iterrows():
    ws8.append([row["季度"]] + [int(row[c]) for c in pivot_q.columns[1:]])

style_header(ws8, HEADER_FILL_GREEN)
style_data_cells(ws8)
auto_column_width(ws8)

chart8 = BarChart()
chart8.type = "col"
chart8.grouping = "stacked"
chart8.title = "各區域季度營收堆疊圖"
chart8.x_axis.title = "季度"
chart8.y_axis.title = "營收（元）"
chart8.y_axis.numFmt = '#,##0'
chart8.width = 24
chart8.height = 15
chart8.style = 10

data_ref8 = Reference(ws8, min_col=2, max_col=ws8.max_column,
                      min_row=1, max_row=ws8.max_row)
cats_ref8 = Reference(ws8, min_col=1, min_row=2, max_row=ws8.max_row)
chart8.add_data(data_ref8, titles_from_data=True)
chart8.set_categories(cats_ref8)

for i, series in enumerate(chart8.series):
    if i < len(CHART_COLORS):
        series.graphicalProperties.solidFill = CHART_COLORS[i]

ws8.add_chart(chart8, "A8")

wb8.save(os.path.join(OUTPUT_DIR, "08_quarterly_stacked_bar.xlsx"))
print("  ✅ 08_quarterly_stacked_bar.xlsx\n")


# ══════════════════════════════════════════════════════════════
# 總覽報表：所有圖表整合到一個檔案
# ══════════════════════════════════════════════════════════════
print("📊 產生總覽報表：chart_dashboard.xlsx...")

wb_all = Workbook()

# --- Sheet 1：月度趨勢 ---
ws_t = wb_all.active
ws_t.title = "月度趨勢"
headers_t = ["月份"] + list(pivot_monthly.columns[1:])
ws_t.append(headers_t)
for _, row in pivot_monthly.iterrows():
    ws_t.append([row["年月"]] + [int(row[r]) for r in pivot_monthly.columns[1:]])
style_header(ws_t)
style_data_cells(ws_t)
auto_column_width(ws_t)

c_t = LineChart()
c_t.title = "2025 年各區域月度營收趨勢"
c_t.x_axis.title = "月份"
c_t.y_axis.title = "營收（元）"
c_t.y_axis.numFmt = '#,##0'
c_t.width = 28
c_t.height = 15
c_t.style = 10
d_t = Reference(ws_t, min_col=2, max_col=ws_t.max_column, min_row=1, max_row=ws_t.max_row)
cat_t = Reference(ws_t, min_col=1, min_row=2, max_row=ws_t.max_row)
c_t.add_data(d_t, titles_from_data=True)
c_t.set_categories(cat_t)
for i, s in enumerate(c_t.series):
    s.graphicalProperties.line.width = 25000
    if i < len(CHART_COLORS):
        s.graphicalProperties.line.solidFill = CHART_COLORS[i]
ws_t.add_chart(c_t, "A15")

# --- Sheet 2：市占率 ---
ws_p = wb_all.create_sheet("市占率")
ws_p.append(["品牌", "市占率(%)", "營收(萬元)"])
for _, row in df_share.iterrows():
    ws_p.append([row["品牌"], row["市占率(%)"], row["營收(萬元)"]])
style_header(ws_p, HEADER_FILL_ORANGE)
style_data_cells(ws_p)
auto_column_width(ws_p)

pie_all = PieChart()
pie_all.title = "市場份額分佈"
pie_all.width = 18
pie_all.height = 14
pie_all.style = 10
pd_all = Reference(ws_p, min_col=2, min_row=1, max_row=ws_p.max_row)
pc_all = Reference(ws_p, min_col=1, min_row=2, max_row=ws_p.max_row)
pie_all.add_data(pd_all, titles_from_data=True)
pie_all.set_categories(pc_all)
pie_all.dataLabels = DataLabelList()
pie_all.dataLabels.showPercent = True
pie_all.dataLabels.showCatName = True
for i, color in enumerate(pie_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie_all.series[0].data_points.append(pt)
ws_p.add_chart(pie_all, "A10")

# --- Sheet 3：預算對比 ---
ws_b = wb_all.create_sheet("預算對比")
ws_b.append(["月份", "預算金額", "實際金額", "差異率(%)"])
for _, row in df_budget.iterrows():
    ws_b.append([row["月份"], int(row["預算金額"]), int(row["實際金額"]), row["差異率(%)"]])
style_header(ws_b, HEADER_FILL_GREEN)
style_data_cells(ws_b)
auto_column_width(ws_b)

bar_all = BarChart()
bar_all.type = "col"
bar_all.grouping = "clustered"
bar_all.title = "月度預算 vs 實際金額"
bar_all.y_axis.title = "金額（元）"
bar_all.y_axis.numFmt = '#,##0'
bar_all.width = 28
bar_all.height = 15
bar_all.style = 10
bd_all = Reference(ws_b, min_col=2, max_col=3, min_row=1, max_row=ws_b.max_row)
bc_all = Reference(ws_b, min_col=1, min_row=2, max_row=ws_b.max_row)
bar_all.add_data(bd_all, titles_from_data=True)
bar_all.set_categories(bc_all)
bar_all.series[0].graphicalProperties.solidFill = "B4C6E7"
bar_all.series[1].graphicalProperties.solidFill = "2F5496"

line_all = LineChart()
line_all.y_axis.title = "差異率（%）"
ld_all = Reference(ws_b, min_col=4, max_col=4, min_row=1, max_row=ws_b.max_row)
line_all.add_data(ld_all, titles_from_data=True)
line_all.series[0].graphicalProperties.line.solidFill = "C00000"
line_all.series[0].graphicalProperties.line.width = 25000
line_all.y_axis.axId = 200
bar_all.y_axis.crosses = "min"
bar_all += line_all
ws_b.add_chart(bar_all, "A16")

# --- Sheet 4：績效排名 ---
ws_r = wb_all.create_sheet("績效排名")
ws_r.append(["業務員", "實際業績", "達成率(%)"])
for _, row in df_perf_sorted.iterrows():
    ws_r.append([row["業務員"], int(row["實際業績"]), row["達成率"]])
style_header(ws_r)
style_data_cells(ws_r)
auto_column_width(ws_r)

bar_r = BarChart()
bar_r.type = "bar"
bar_r.grouping = "clustered"
bar_r.title = "業務員年度業績排名"
bar_r.y_axis.numFmt = '#,##0'
bar_r.width = 26
bar_r.height = 16
bar_r.style = 10
dr = Reference(ws_r, min_col=2, max_col=2, min_row=1, max_row=ws_r.max_row)
cr = Reference(ws_r, min_col=1, min_row=2, max_row=ws_r.max_row)
bar_r.add_data(dr, titles_from_data=True)
bar_r.set_categories(cr)
bar_r.series[0].graphicalProperties.solidFill = "2F5496"
ws_r.add_chart(bar_r, "E1")

wb_all.save(os.path.join(OUTPUT_DIR, "chart_dashboard.xlsx"))
print("  ✅ chart_dashboard.xlsx\n")

print("=" * 50)
print("🎉 所有圖表報表已產生於 output/ 資料夾")
print(f"   共 9 個 Excel 檔案（8 個獨立圖表 + 1 個總覽儀表板）")
print("=" * 50)
