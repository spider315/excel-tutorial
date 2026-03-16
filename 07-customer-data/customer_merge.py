#!/usr/bin/env python3
"""
07-customer-data: 客戶資料清洗與合併 — 主處理腳本
讀取三份客戶名單，清理並合併為單一主檔
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# ══════════════════════════════════════════════════════════════════════════════
# 路徑設定
# ══════════════════════════════════════════════════════════════════════════════
RAW_DIR = Path(__file__).parent / "raw"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# 樣式定義
# ══════════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DUP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def style_header(ws, fill=HEADER_FILL):
    """套用標題列樣式"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """自動調整欄寬"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


# ══════════════════════════════════════════════════════════════════════════════
# 電話號碼標準化
# ══════════════════════════════════════════════════════════════════════════════
def standardize_phone(phone):
    """標準化電話號碼為 0912345678 格式"""
    if pd.isna(phone):
        return ""

    phone = str(phone).strip()

    # 中文數字轉換
    chinese_nums = {"零": "0", "一": "1", "二": "2", "三": "3", "四": "4",
                    "五": "5", "六": "6", "七": "7", "八": "8", "九": "9"}
    for cn, num in chinese_nums.items():
        phone = phone.replace(cn, num)

    # 移除所有非數字字元
    phone = re.sub(r"[^\d]", "", phone)

    # 處理國際碼
    if phone.startswith("886"):
        phone = "0" + phone[3:]
    elif phone.startswith("09") and len(phone) == 10:
        pass  # 已經是正確格式
    elif phone.startswith("9") and len(phone) == 9:
        phone = "0" + phone

    # 驗證格式
    if len(phone) == 10 and phone.startswith("09"):
        return phone
    else:
        return phone  # 回傳原值，但標記為需檢查


# ══════════════════════════════════════════════════════════════════════════════
# Email 標準化
# ══════════════════════════════════════════════════════════════════════════════
def standardize_email(email):
    """標準化 Email 為小寫並驗證格式"""
    if pd.isna(email):
        return ""

    email = str(email).strip().lower()

    # 簡單驗證
    if "@" in email and "." in email.split("@")[-1]:
        # 檢查 @ 前後都有字元
        parts = email.split("@")
        if len(parts) == 2 and parts[0] and parts[1]:
            return email

    return ""  # 無效的 email


# ══════════════════════════════════════════════════════════════════════════════
# Task 1: 讀取並標準化各來源資料
# ══════════════════════════════════════════════════════════════════════════════
def load_and_standardize():
    """讀取並標準化三份客戶名單"""
    print("\n📋 Task 1: 讀取並標準化資料")
    print("-" * 40)

    cleaning_log = []

    # ===== CRM 資料 =====
    crm_df = pd.read_excel(RAW_DIR / "crm_customers.xlsx", dtype=str)
    crm_df = crm_df.rename(columns={
        "客戶編號": "原始編號",
        "客戶姓名": "姓名",
        "連絡電話": "電話",
        "電子郵件": "Email",
        "公司名稱": "公司",
        "產業類別": "產業",
        "地址": "地址",
        "來源管道": "來源",
        "建檔日期": "建檔日期",
    })
    crm_df["資料來源"] = "CRM"

    # ===== 業務 Excel =====
    sales_df = pd.read_excel(RAW_DIR / "sales_customers.xlsx", dtype=str)
    sales_df = sales_df.rename(columns={
        "電話": "電話",
        "Email": "Email",
        "公司": "公司",
        "業務負責人": "業務",
        "備註": "備註",
    })
    sales_df["原始編號"] = [f"SALES{i+1:04d}" for i in range(len(sales_df))]
    sales_df["資料來源"] = "業務Excel"

    # ===== 網站註冊 =====
    web_df = pd.read_excel(RAW_DIR / "web_registrations.xlsx", dtype=str)
    web_df = web_df.rename(columns={
        "註冊ID": "原始編號",
        "full_name": "姓名",
        "phone_number": "電話",
        "email_address": "Email",
        "registration_date": "建檔日期",
        "newsletter_subscribed": "訂閱電子報",
    })
    web_df["資料來源"] = "網站註冊"

    # ===== 標準化電話 =====
    for df, source in [(crm_df, "CRM"), (sales_df, "業務Excel"), (web_df, "網站註冊")]:
        for idx, val in df["電話"].items():
            new_val = standardize_phone(val)
            if new_val != str(val).strip():
                cleaning_log.append({
                    "資料來源": source, "欄位": "電話", "列號": idx + 2,
                    "原始值": val, "修正值": new_val, "原因": "電話格式標準化"
                })
                df.at[idx, "電話"] = new_val

    # ===== 標準化 Email =====
    for df, source in [(crm_df, "CRM"), (sales_df, "業務Excel"), (web_df, "網站註冊")]:
        for idx, val in df["Email"].items():
            new_val = standardize_email(val)
            if new_val != str(val).strip().lower() and new_val:
                cleaning_log.append({
                    "資料來源": source, "欄位": "Email", "列號": idx + 2,
                    "原始值": val, "修正值": new_val, "原因": "Email格式標準化"
                })
                df.at[idx, "Email"] = new_val
            elif not new_val and val:
                cleaning_log.append({
                    "資料來源": source, "欄位": "Email", "列號": idx + 2,
                    "原始值": val, "修正值": "(無效)", "原因": "Email格式錯誤"
                })
                df.at[idx, "Email"] = ""

    # ===== 姓名去空白 =====
    for df, source in [(crm_df, "CRM"), (sales_df, "業務Excel"), (web_df, "網站註冊")]:
        for idx, val in df["姓名"].items():
            new_val = str(val).strip()
            if new_val != val:
                cleaning_log.append({
                    "資料來源": source, "欄位": "姓名", "列號": idx + 2,
                    "原始值": repr(val), "修正值": new_val, "原因": "移除空白"
                })
                df.at[idx, "姓名"] = new_val

    # 輸出清理日誌
    log_df = pd.DataFrame(cleaning_log)
    log_df.to_excel(OUTPUT_DIR / "cleaning_log.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaning_log.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaning_log.xlsx")

    print(f"  ✅ CRM 資料：{len(crm_df)} 筆")
    print(f"  ✅ 業務 Excel：{len(sales_df)} 筆")
    print(f"  ✅ 網站註冊：{len(web_df)} 筆")
    print(f"  ✅ 清理修正：{len(cleaning_log)} 處")
    print(f"  📄 輸出：cleaning_log.xlsx")

    return crm_df, sales_df, web_df, cleaning_log


# ══════════════════════════════════════════════════════════════════════════════
# Task 2: 合併並去重
# ══════════════════════════════════════════════════════════════════════════════
def merge_and_deduplicate(crm_df, sales_df, web_df):
    """合併三份名單並去除重複"""
    print("\n📋 Task 2: 合併並去重")
    print("-" * 40)

    # 統一欄位
    common_cols = ["原始編號", "姓名", "電話", "Email", "公司", "資料來源"]

    # 確保所有 DataFrame 都有這些欄位
    for df in [crm_df, sales_df, web_df]:
        for col in common_cols:
            if col not in df.columns:
                df[col] = ""

    # 選取共同欄位
    crm_selected = crm_df[common_cols].copy()
    sales_selected = sales_df[common_cols].copy()
    web_selected = web_df[common_cols].copy()

    # 合併
    merged_df = pd.concat([crm_selected, sales_selected, web_selected], ignore_index=True)
    print(f"  合併後總筆數：{len(merged_df)} 筆")

    # 找出重複（依電話或 Email）
    merged_df["電話_標準"] = merged_df["電話"].apply(lambda x: x if len(str(x)) == 10 else "")
    merged_df["Email_標準"] = merged_df["Email"].apply(lambda x: x if "@" in str(x) else "")

    # 標記重複
    merged_df["重複標記"] = ""

    # 依電話找重複
    phone_counts = merged_df[merged_df["電話_標準"] != ""]["電話_標準"].value_counts()
    dup_phones = phone_counts[phone_counts > 1].index.tolist()

    for phone in dup_phones:
        mask = merged_df["電話_標準"] == phone
        merged_df.loc[mask, "重複標記"] = "電話重複"

    # 依 Email 找重複
    email_counts = merged_df[merged_df["Email_標準"] != ""]["Email_標準"].value_counts()
    dup_emails = email_counts[email_counts > 1].index.tolist()

    for email in dup_emails:
        mask = merged_df["Email_標準"] == email
        if merged_df.loc[mask, "重複標記"].eq("").all():
            merged_df.loc[mask, "重複標記"] = "Email重複"
        else:
            merged_df.loc[mask, "重複標記"] = "電話+Email重複"

    # 分離重複與不重複
    duplicates = merged_df[merged_df["重複標記"] != ""].copy()
    unique = merged_df[merged_df["重複標記"] == ""].copy()

    # 處理重複：保留第一筆
    deduped = merged_df.drop_duplicates(subset=["電話_標準"], keep="first")
    deduped = deduped.drop_duplicates(subset=["Email_標準"], keep="first")

    # 移除工作欄位
    final_cols = ["原始編號", "姓名", "電話", "Email", "公司", "資料來源"]
    deduped = deduped[final_cols]

    print(f"  重複筆數：{len(duplicates)} 筆")
    print(f"  去重後筆數：{len(deduped)} 筆")

    return deduped, duplicates


# ══════════════════════════════════════════════════════════════════════════════
# Task 3: 輸出報表
# ══════════════════════════════════════════════════════════════════════════════
def generate_reports(deduped, duplicates):
    """產生合併報表"""
    print("\n📋 Task 3: 輸出報表")
    print("-" * 40)

    # 主檔
    deduped.to_excel(OUTPUT_DIR / "customer_master.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "customer_master.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "customer_master.xlsx")
    print(f"  📄 輸出：customer_master.xlsx ({len(deduped)} 筆)")

    # 重複清單
    if len(duplicates) > 0:
        duplicates.to_excel(OUTPUT_DIR / "duplicate_records.xlsx", index=False)
        wb = load_workbook(OUTPUT_DIR / "duplicate_records.xlsx")
        ws = wb.active
        style_header(ws)
        auto_column_width(ws)

        # 標記重複列
        dup_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "重複標記":
                dup_col = idx
                break

        if dup_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=dup_col).value:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = DUP_FILL

        wb.save(OUTPUT_DIR / "duplicate_records.xlsx")
        print(f"  📄 輸出：duplicate_records.xlsx ({len(duplicates)} 筆)")

    # 來源統計
    source_summary = deduped["資料來源"].value_counts().reset_index()
    source_summary.columns = ["資料來源", "客戶數"]
    source_summary.to_excel(OUTPUT_DIR / "source_summary.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "source_summary.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "source_summary.xlsx")
    print(f"  📄 輸出：source_summary.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("📊 客戶資料清洗與合併 — 主處理腳本")
    print("=" * 60)

    # Task 1: 讀取並標準化
    crm_df, sales_df, web_df, cleaning_log = load_and_standardize()

    # Task 2: 合併並去重
    deduped, duplicates = merge_and_deduplicate(crm_df, sales_df, web_df)

    # Task 3: 輸出報表
    generate_reports(deduped, duplicates)

    print("\n" + "=" * 60)
    print("✅ 客戶資料處理完成！")
    print(f"📁 輸出目錄: {OUTPUT_DIR}")
    print("=" * 60)
