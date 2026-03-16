#!/usr/bin/env python3
"""
09-survey: 問卷調查統計分析 — 主處理腳本
讀取問卷結果，產生統計報表與圖表
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, RadarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
import re

# ══════════════════════════════════════════════════════════════════════════════
# 路徑設定
# ══════════════════════════════════════════════════════════════════════════════
RAW_DIR = Path(__file__).parent / "raw"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# 樣式定義
# ══════════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MID_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 題目對照
QUESTIONS = {
    "Q1": "整體工作滿意度",
    "Q2": "對直屬主管的滿意度",
    "Q3": "團隊合作氛圍",
    "Q4": "薪資福利滿意度",
    "Q5": "工作與生活平衡",
    "Q6": "職涯發展機會",
    "Q7": "公司制度與流程",
    "Q8": "教育訓練資源",
    "Q9": "辦公環境設備",
    "Q10": "對公司未來的信心",
}


def style_header(ws, fill=HEADER_FILL):
    """套用標題列樣式"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """自動調整欄寬"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


# ══════════════════════════════════════════════════════════════════════════════
# Task 1: 資料清理
# ══════════════════════════════════════════════════════════════════════════════
def clean_survey_data():
    """清理問卷資料"""
    print("\n📋 Task 1: 資料清理")
    print("-" * 40)

    df = pd.read_excel(RAW_DIR / "survey_responses.xlsx")
    cleaning_log = []

    # 1. 分數範圍修正（1-5）
    q_cols = [col for col in df.columns if col.startswith("Q")]
    for col in q_cols:
        for idx, val in df[col].items():
            if pd.notna(val):
                if val < 1:
                    cleaning_log.append({
                        "欄位": col, "列號": idx + 2,
                        "原始值": val, "修正值": 1, "原因": "分數低於最小值"
                    })
                    df.at[idx, col] = 1
                elif val > 5:
                    cleaning_log.append({
                        "欄位": col, "列號": idx + 2,
                        "原始值": val, "修正值": 5, "原因": "分數超過最大值"
                    })
                    df.at[idx, col] = 5

    # 2. 空值填入中位數
    for col in q_cols:
        median = df[col].median()
        for idx, val in df[col].items():
            if pd.isna(val):
                cleaning_log.append({
                    "欄位": col, "列號": idx + 2,
                    "原始值": "空值", "修正值": median, "原因": "空值以中位數填補"
                })
                df.at[idx, col] = median

    # 3. 日期格式標準化
    def fix_date(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        patterns = [
            (r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", r"\1/\2/\3"),
            (r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})", r"\3/\2/\1"),
        ]
        for pattern, replacement in patterns:
            if re.match(pattern, val):
                val = re.sub(pattern, replacement, val)
                break
        return val

    for idx, val in df["填答日期"].items():
        new_val = fix_date(val)
        if new_val != str(val):
            cleaning_log.append({
                "欄位": "填答日期", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "日期格式標準化"
            })
            df.at[idx, "填答日期"] = new_val

    # 4. 部門名稱標準化
    def fix_dept(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        val = re.sub(r"\s+", "", val)
        if val and not val.endswith("部"):
            val = val + "部"
        return val

    for idx, val in df["部門"].items():
        new_val = fix_dept(val)
        if new_val != str(val):
            cleaning_log.append({
                "欄位": "部門", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "部門名稱標準化"
            })
            df.at[idx, "部門"] = new_val

    # 5. 職級標準化
    level_mapping = {
        "一般": "一般職員",
        "經裡": "經理",
    }
    for idx, val in df["職級"].items():
        if val in level_mapping:
            new_val = level_mapping[val]
            cleaning_log.append({
                "欄位": "職級", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "職級名稱修正"
            })
            df.at[idx, "職級"] = new_val

    # 輸出
    df.to_excel(OUTPUT_DIR / "cleaned_survey.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaned_survey.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaned_survey.xlsx")

    log_df = pd.DataFrame(cleaning_log)
    log_df.to_excel(OUTPUT_DIR / "cleaning_log.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaning_log.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaning_log.xlsx")

    print(f"  ✅ 清理完成：{len(cleaning_log)} 處修正")
    print(f"  📄 輸出：cleaned_survey.xlsx")
    print(f"  📄 輸出：cleaning_log.xlsx")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# Task 2: 部門分析
# ══════════════════════════════════════════════════════════════════════════════
def analyze_by_department(df):
    """依部門分析滿意度"""
    print("\n📋 Task 2: 部門分析")
    print("-" * 40)

    q_cols = [col for col in df.columns if col.startswith("Q")]

    # 部門平均分數
    dept_avg = df.groupby("部門")[q_cols].mean().round(2)
    dept_avg["整體平均"] = dept_avg.mean(axis=1).round(2)

    # 計算排名
    dept_avg["排名"] = dept_avg["整體平均"].rank(ascending=False).astype(int)
    dept_avg = dept_avg.sort_values("整體平均", ascending=False)

    # 重命名欄位
    rename_dict = {q: QUESTIONS[q] for q in q_cols}
    dept_avg_display = dept_avg.rename(columns=rename_dict)

    dept_avg_display.to_excel(OUTPUT_DIR / "department_analysis.xlsx")

    wb = load_workbook(OUTPUT_DIR / "department_analysis.xlsx")
    ws = wb.active
    style_header(ws)
    auto_column_width(ws)

    # 條件格式：整體平均分數
    avg_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "整體平均":
            avg_col = idx
            break

    if avg_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=avg_col)
            if cell.value:
                if cell.value < 3.5:
                    cell.fill = LOW_FILL
                elif cell.value < 4.0:
                    cell.fill = MID_FILL
                else:
                    cell.fill = HIGH_FILL

    wb.save(OUTPUT_DIR / "department_analysis.xlsx")

    print(f"  📄 輸出：department_analysis.xlsx")

    return dept_avg


# ══════════════════════════════════════════════════════════════════════════════
# Task 3: 題目分析
# ══════════════════════════════════════════════════════════════════════════════
def analyze_by_question(df):
    """依題目分析滿意度"""
    print("\n📋 Task 3: 題目分析")
    print("-" * 40)

    q_cols = [col for col in df.columns if col.startswith("Q")]

    rows = []
    for q in q_cols:
        scores = df[q]
        rows.append({
            "題目代碼": q,
            "題目內容": QUESTIONS[q],
            "平均分數": round(scores.mean(), 2),
            "標準差": round(scores.std(), 2),
            "最高分": scores.max(),
            "最低分": scores.min(),
            "滿意比例": f"{(scores >= 4).sum() / len(scores) * 100:.1f}%",
            "不滿意比例": f"{(scores <= 2).sum() / len(scores) * 100:.1f}%",
        })

    q_analysis = pd.DataFrame(rows)
    q_analysis = q_analysis.sort_values("平均分數", ascending=False)

    q_analysis.to_excel(OUTPUT_DIR / "question_analysis.xlsx", index=False)

    wb = load_workbook(OUTPUT_DIR / "question_analysis.xlsx")
    ws = wb.active
    style_header(ws)
    auto_column_width(ws)

    # 條件格式：平均分數
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)  # 平均分數欄
        if cell.value:
            if cell.value < 3.5:
                cell.fill = LOW_FILL
            elif cell.value < 4.0:
                cell.fill = MID_FILL
            else:
                cell.fill = HIGH_FILL

    wb.save(OUTPUT_DIR / "question_analysis.xlsx")

    print(f"  📄 輸出：question_analysis.xlsx")

    return q_analysis


# ══════════════════════════════════════════════════════════════════════════════
# Task 4: 交叉分析
# ══════════════════════════════════════════════════════════════════════════════
def cross_analysis(df):
    """職級 × 年資交叉分析"""
    print("\n📋 Task 4: 交叉分析")
    print("-" * 40)

    q_cols = [col for col in df.columns if col.startswith("Q")]

    # 計算整體滿意度平均（所有題目的平均）
    df["整體滿意度"] = df[q_cols].mean(axis=1)

    # 職級 × 年資 樞紐表
    pivot = df.pivot_table(
        values="整體滿意度",
        index="職級",
        columns="年資",
        aggfunc="mean"
    ).round(2)

    # 排序
    level_order = ["一般職員", "資深專員", "主任", "經理", "副總"]
    tenure_order = ["未滿1年", "1-3年", "3-5年", "5-10年", "10年以上"]

    pivot = pivot.reindex(index=[l for l in level_order if l in pivot.index])
    pivot = pivot.reindex(columns=[t for t in tenure_order if t in pivot.columns])

    pivot.to_excel(OUTPUT_DIR / "cross_analysis.xlsx")

    wb = load_workbook(OUTPUT_DIR / "cross_analysis.xlsx")
    ws = wb.active
    style_header(ws)
    auto_column_width(ws)

    # 數值條件格式
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                if cell.value < 3.5:
                    cell.fill = LOW_FILL
                elif cell.value < 4.0:
                    cell.fill = MID_FILL
                else:
                    cell.fill = HIGH_FILL

    wb.save(OUTPUT_DIR / "cross_analysis.xlsx")

    print(f"  📄 輸出：cross_analysis.xlsx")

    return pivot


# ══════════════════════════════════════════════════════════════════════════════
# Task 5: 圖表儀表板
# ══════════════════════════════════════════════════════════════════════════════
def create_dashboard(dept_avg, q_analysis, df):
    """建立圖表儀表板"""
    print("\n📋 Task 5: 圖表儀表板")
    print("-" * 40)

    q_cols = [col for col in df.columns if col.startswith("Q")]

    with pd.ExcelWriter(OUTPUT_DIR / "survey_dashboard.xlsx", engine="openpyxl") as writer:
        # Sheet 1: 部門平均
        dept_display = dept_avg.reset_index()[["部門", "整體平均"]]
        dept_display.to_excel(writer, sheet_name="部門比較", index=False)

        # Sheet 2: 題目平均
        q_display = q_analysis[["題目內容", "平均分數"]]
        q_display.to_excel(writer, sheet_name="題目比較", index=False)

        # Sheet 3: 整體統計
        stats = pd.DataFrame({
            "指標": ["總填答人數", "整體平均分數", "最高分題目", "最低分題目", "最高分部門", "最低分部門"],
            "數值": [
                len(df),
                round(df[q_cols].mean().mean(), 2),
                q_analysis.iloc[0]["題目內容"],
                q_analysis.iloc[-1]["題目內容"],
                dept_avg.index[0],
                dept_avg.index[-1],
            ]
        })
        stats.to_excel(writer, sheet_name="整體統計", index=False)

    wb = load_workbook(OUTPUT_DIR / "survey_dashboard.xlsx")

    # 部門比較圖表
    ws = wb["部門比較"]
    style_header(ws)
    auto_column_width(ws)

    chart = BarChart()
    chart.type = "col"
    chart.title = "各部門整體滿意度"
    chart.y_axis.title = "平均分數"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 5

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10

    ws.add_chart(chart, "D2")

    # 題目比較圖表
    ws = wb["題目比較"]
    style_header(ws)
    auto_column_width(ws)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "各題目平均分數"
    chart.x_axis.title = "平均分數"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 5

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 12

    ws.add_chart(chart, "D2")

    # 整體統計樣式
    ws = wb["整體統計"]
    style_header(ws)
    auto_column_width(ws)

    wb.save(OUTPUT_DIR / "survey_dashboard.xlsx")

    print(f"  📄 輸出：survey_dashboard.xlsx (含圖表)")


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("📊 問卷調查統計分析 — 主處理腳本")
    print("=" * 60)

    # Task 1: 資料清理
    df = clean_survey_data()

    # Task 2: 部門分析
    dept_avg = analyze_by_department(df)

    # Task 3: 題目分析
    q_analysis = analyze_by_question(df)

    # Task 4: 交叉分析
    cross_analysis(df)

    # Task 5: 圖表儀表板
    create_dashboard(dept_avg, q_analysis, df)

    print("\n" + "=" * 60)
    print("✅ 問卷分析處理完成！")
    print(f"📁 輸出目錄: {OUTPUT_DIR}")
    print("=" * 60)
