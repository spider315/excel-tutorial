#!/usr/bin/env python3
"""
06-gantt: 專案排程甘特圖 — 主處理腳本
讀取專案任務清單，產生甘特圖樣式的 Excel 報表
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
import re

# ══════════════════════════════════════════════════════════════════════════════
# 路徑設定
# ══════════════════════════════════════════════════════════════════════════════
RAW_DIR = Path(__file__).parent / "raw"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# 樣式定義
# ══════════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
PHASE_FILL = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")

# 狀態顏色
STATUS_COLORS = {
    "已完成": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "進行中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "未開始": PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),
    "延遲": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

# 甘特條顏色
GANTT_COMPLETE = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
GANTT_PROGRESS = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
GANTT_PLANNED = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
GANTT_DELAYED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def style_header(ws, row=1, fill=HEADER_FILL):
    """套用標題列樣式"""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_column_width(ws, min_width=8, max_width=25):
    """自動調整欄寬"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max(min(max_length + 2, max_width), min_width)


# ══════════════════════════════════════════════════════════════════════════════
# Task 1: 資料清理
# ══════════════════════════════════════════════════════════════════════════════
def clean_task_data():
    """清理專案任務資料"""
    print("\n📋 Task 1: 資料清理")
    print("-" * 40)

    df = pd.read_excel(RAW_DIR / "project_tasks.xlsx", dtype=str)
    cleaning_log = []

    # 1. 日期格式統一
    def standardize_date(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        patterns = [
            (r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", r"\1/\2/\3"),
            (r"(\d{2})[/\-.](\d{1,2})[/\-.](\d{1,2})", r"20\1/\2/\3"),
        ]
        for pattern, replacement in patterns:
            if re.match(pattern, val):
                val = re.sub(pattern, replacement, val)
                break
        return val

    for col in ["開始日期", "結束日期"]:
        for idx, val in df[col].items():
            new_val = standardize_date(val)
            if new_val != val:
                cleaning_log.append({
                    "欄位": col, "列號": idx + 2,
                    "原始值": val, "修正值": new_val, "原因": "日期格式標準化"
                })
                df.at[idx, col] = new_val

    # 2. 負責人名字去除空白
    for idx, val in df["負責人"].items():
        new_val = str(val).strip()
        if new_val != val:
            cleaning_log.append({
                "欄位": "負責人", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "移除多餘空白"
            })
            df.at[idx, "負責人"] = new_val

    # 3. 進度與狀態一致性檢查
    df["進度(%)"] = pd.to_numeric(df["進度(%)"], errors="coerce").fillna(0).astype(int)
    for idx, row in df.iterrows():
        progress = row["進度(%)"]
        status = row["狀態"]
        if progress == 100 and status != "已完成":
            cleaning_log.append({
                "欄位": "狀態", "列號": idx + 2,
                "原始值": status, "修正值": "已完成", "原因": "進度100%但狀態不一致"
            })
            df.at[idx, "狀態"] = "已完成"
        elif progress > 0 and progress < 100 and status == "未開始":
            cleaning_log.append({
                "欄位": "狀態", "列號": idx + 2,
                "原始值": status, "修正值": "進行中", "原因": "有進度但狀態為未開始"
            })
            df.at[idx, "狀態"] = "進行中"

    # 4. 修正開始日期 > 結束日期的問題
    for idx, row in df.iterrows():
        try:
            start = datetime.strptime(row["開始日期"], "%Y/%m/%d")
            end = datetime.strptime(row["結束日期"], "%Y/%m/%d")
            if start > end:
                # 交換日期
                cleaning_log.append({
                    "欄位": "日期", "列號": idx + 2,
                    "原始值": f"{row['開始日期']} ~ {row['結束日期']}",
                    "修正值": f"{row['結束日期']} ~ {row['開始日期']}",
                    "原因": "開始日期晚於結束日期，已交換"
                })
                df.at[idx, "開始日期"] = row["結束日期"]
                df.at[idx, "結束日期"] = row["開始日期"]
        except:
            pass

    # 5. 重新計算工期
    for idx, row in df.iterrows():
        try:
            start = datetime.strptime(row["開始日期"], "%Y/%m/%d")
            end = datetime.strptime(row["結束日期"], "%Y/%m/%d")
            correct_duration = (end - start).days
            if int(row["工期(天)"]) != correct_duration:
                cleaning_log.append({
                    "欄位": "工期(天)", "列號": idx + 2,
                    "原始值": row["工期(天)"], "修正值": correct_duration, "原因": "工期與日期範圍不符"
                })
                df.at[idx, "工期(天)"] = correct_duration
        except:
            pass

    # 輸出清理後資料
    df.to_excel(OUTPUT_DIR / "cleaned_tasks.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaned_tasks.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaned_tasks.xlsx")

    # 輸出清理日誌
    log_df = pd.DataFrame(cleaning_log)
    log_df.to_excel(OUTPUT_DIR / "cleaning_log.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaning_log.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaning_log.xlsx")

    print(f"  ✅ 清理完成：{len(cleaning_log)} 處修正")
    print(f"  📄 輸出：cleaned_tasks.xlsx")
    print(f"  📄 輸出：cleaning_log.xlsx")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# Task 2: 甘特圖產生
# ══════════════════════════════════════════════════════════════════════════════
def generate_gantt_chart(task_df):
    """產生甘特圖樣式的 Excel"""
    print("\n📋 Task 2: 甘特圖產生")
    print("-" * 40)

    wb = Workbook()
    ws = wb.active
    ws.title = "甘特圖"

    # 取得日期範圍
    task_df["開始日期_dt"] = pd.to_datetime(task_df["開始日期"], format="%Y/%m/%d")
    task_df["結束日期_dt"] = pd.to_datetime(task_df["結束日期"], format="%Y/%m/%d")

    min_date = task_df["開始日期_dt"].min()
    max_date = task_df["結束日期_dt"].max()

    # 建立日期欄位（以週為單位）
    date_columns = []
    current_date = min_date
    while current_date <= max_date + timedelta(days=7):
        date_columns.append(current_date)
        current_date += timedelta(days=7)

    # 標題列
    headers = ["任務編號", "任務名稱", "負責人", "開始日期", "結束日期", "進度", "狀態"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 日期標題
    for col, date in enumerate(date_columns, len(headers) + 1):
        ws.cell(row=1, column=col, value=date.strftime("%m/%d"))

    style_header(ws)

    # 填入任務資料
    today = datetime.now()
    for row, (_, task) in enumerate(task_df.iterrows(), 2):
        ws.cell(row=row, column=1, value=task["任務編號"])
        ws.cell(row=row, column=2, value=task["任務名稱"])
        ws.cell(row=row, column=3, value=task["負責人"])
        ws.cell(row=row, column=4, value=task["開始日期"])
        ws.cell(row=row, column=5, value=task["結束日期"])
        ws.cell(row=row, column=6, value=f"{task['進度(%)']}%")
        ws.cell(row=row, column=7, value=task["狀態"])

        # 狀態顏色
        status = task["狀態"]
        if status in STATUS_COLORS:
            ws.cell(row=row, column=7).fill = STATUS_COLORS[status]

        # 甘特條
        task_start = task["開始日期_dt"]
        task_end = task["結束日期_dt"]
        progress = int(task["進度(%)"])

        for col, week_start in enumerate(date_columns, len(headers) + 1):
            week_end = week_start + timedelta(days=6)

            # 檢查這週是否在任務範圍內
            if task_end >= week_start and task_start <= week_end:
                cell = ws.cell(row=row, column=col)

                # 判斷顏色
                if status == "已完成":
                    cell.fill = GANTT_COMPLETE
                elif status == "延遲":
                    cell.fill = GANTT_DELAYED
                elif progress > 0:
                    # 計算進度對應的週數
                    total_weeks = (task_end - task_start).days / 7
                    progress_weeks = total_weeks * (progress / 100)
                    current_week = (week_start - task_start).days / 7

                    if current_week < progress_weeks:
                        cell.fill = GANTT_PROGRESS
                    else:
                        cell.fill = GANTT_PLANNED
                else:
                    cell.fill = GANTT_PLANNED

    # 調整欄寬
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8

    for col in range(len(headers) + 1, len(headers) + len(date_columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6

    wb.save(OUTPUT_DIR / "gantt_chart.xlsx")
    print(f"  ✅ 甘特圖產生完成")
    print(f"  📄 輸出：gantt_chart.xlsx")

    return wb


# ══════════════════════════════════════════════════════════════════════════════
# Task 3: 進度追蹤報表
# ══════════════════════════════════════════════════════════════════════════════
def generate_progress_report(task_df):
    """產生進度追蹤報表"""
    print("\n📋 Task 3: 進度追蹤報表")
    print("-" * 40)

    # 階段彙總
    phase_summary = task_df.groupby(["階段代碼", "階段名稱"]).agg({
        "任務編號": "count",
        "進度(%)": "mean",
        "工期(天)": "sum"
    }).reset_index()
    phase_summary.columns = ["階段代碼", "階段名稱", "任務數", "平均進度(%)", "總工期(天)"]
    phase_summary["平均進度(%)"] = phase_summary["平均進度(%)"].round(1)

    # 負責人工作量
    member_workload = task_df.groupby("負責人").agg({
        "任務編號": "count",
        "工期(天)": "sum",
        "進度(%)": "mean"
    }).reset_index()
    member_workload.columns = ["負責人", "任務數", "總工期(天)", "平均進度(%)"]
    member_workload["平均進度(%)"] = member_workload["平均進度(%)"].round(1)

    # 狀態統計
    status_summary = task_df["狀態"].value_counts().reset_index()
    status_summary.columns = ["狀態", "數量"]

    # 輸出
    with pd.ExcelWriter(OUTPUT_DIR / "progress_report.xlsx", engine="openpyxl") as writer:
        phase_summary.to_excel(writer, sheet_name="階段進度", index=False)
        member_workload.to_excel(writer, sheet_name="人員工作量", index=False)
        status_summary.to_excel(writer, sheet_name="狀態統計", index=False)

    wb = load_workbook(OUTPUT_DIR / "progress_report.xlsx")
    for ws in wb.worksheets:
        style_header(ws)
        auto_column_width(ws)
    wb.save(OUTPUT_DIR / "progress_report.xlsx")

    print(f"  ✅ 進度報表產生完成")
    print(f"  📄 輸出：progress_report.xlsx (3 個工作表)")

    return phase_summary


# ══════════════════════════════════════════════════════════════════════════════
# Task 4: 里程碑追蹤
# ══════════════════════════════════════════════════════════════════════════════
def track_milestones():
    """追蹤里程碑達成狀態"""
    print("\n📋 Task 4: 里程碑追蹤")
    print("-" * 40)

    milestone_df = pd.read_excel(RAW_DIR / "milestones.xlsx")
    today = datetime.now()

    # 計算剩餘天數
    milestone_df["目標日期_dt"] = pd.to_datetime(milestone_df["目標日期"], format="%Y/%m/%d")
    milestone_df["剩餘天數"] = (milestone_df["目標日期_dt"] - today).dt.days
    milestone_df = milestone_df.drop(columns=["目標日期_dt"])

    # 判斷風險
    def assess_risk(row):
        if row["達成狀態"] == "已達成":
            return "無"
        elif row["剩餘天數"] < 0:
            return "高（已逾期）"
        elif row["剩餘天數"] < 7:
            return "中（7天內）"
        else:
            return "低"

    milestone_df["風險等級"] = milestone_df.apply(assess_risk, axis=1)

    milestone_df.to_excel(OUTPUT_DIR / "milestone_tracking.xlsx", index=False)

    wb = load_workbook(OUTPUT_DIR / "milestone_tracking.xlsx")
    ws = wb.active
    style_header(ws)
    auto_column_width(ws)

    # 風險顏色
    risk_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "風險等級":
            risk_col = idx
            break

    if risk_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=risk_col)
            if "高" in str(cell.value):
                cell.fill = STATUS_COLORS["延遲"]
            elif "中" in str(cell.value):
                cell.fill = STATUS_COLORS["進行中"]

    wb.save(OUTPUT_DIR / "milestone_tracking.xlsx")

    print(f"  ✅ 里程碑追蹤完成")
    print(f"  📄 輸出：milestone_tracking.xlsx")

    return milestone_df


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("📊 專案排程甘特圖 — 主處理腳本")
    print("=" * 60)

    # Task 1: 資料清理
    task_df = clean_task_data()

    # Task 2: 甘特圖產生
    generate_gantt_chart(task_df)

    # Task 3: 進度追蹤報表
    generate_progress_report(task_df)

    # Task 4: 里程碑追蹤
    track_milestones()

    print("\n" + "=" * 60)
    print("✅ 所有甘特圖報表處理完成！")
    print(f"📁 輸出目錄: {OUTPUT_DIR}")
    print("=" * 60)
