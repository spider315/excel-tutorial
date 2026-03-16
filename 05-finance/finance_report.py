#!/usr/bin/env python3
"""
05-finance: 財務報表自動化 — 主處理腳本
讀取 raw/ 中的費用明細，產生損益表、資產負債表、預算達成分析
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
import re

# ══════════════════════════════════════════════════════════════════════════════
# 路徑設定
# ══════════════════════════════════════════════════════════════════════════════
RAW_DIR = Path(__file__).parent / "raw"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# 樣式定義
# ══════════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def style_header(ws, fill=HEADER_FILL):
    """套用標題列樣式"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """自動調整欄寬"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


# ══════════════════════════════════════════════════════════════════════════════
# Task 1: 資料清理
# ══════════════════════════════════════════════════════════════════════════════
def clean_expense_data():
    """清理費用明細資料"""
    print("\n📋 Task 1: 資料清理")
    print("-" * 40)

    df = pd.read_excel(RAW_DIR / "expense_detail_2025.xlsx", dtype=str)
    df_orig = df.copy()
    cleaning_log = []

    # 1. 日期格式統一
    def standardize_date(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        # 處理各種格式
        patterns = [
            (r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", r"\1/\2/\3"),  # 2025-01-06 -> 2025/1/6
            (r"(\d{2})[/\-.](\d{1,2})[/\-.](\d{1,2})", r"20\1/\2/\3"),  # 25/02/15 -> 2025/2/15
        ]
        for pattern, replacement in patterns:
            if re.match(pattern, val):
                val = re.sub(pattern, replacement, val)
                break
        return val

    for idx, val in df["日期"].items():
        new_val = standardize_date(val)
        if new_val != val:
            cleaning_log.append({
                "欄位": "日期", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "日期格式標準化"
            })
            df.at[idx, "日期"] = new_val

    # 2. 金額修正（負數轉正數）
    df["借方金額"] = pd.to_numeric(df["借方金額"], errors="coerce").fillna(0)
    for idx, val in enumerate(df["借方金額"]):
        if val < 0:
            new_val = abs(val)
            cleaning_log.append({
                "欄位": "借方金額", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "負數金額轉正"
            })
            df.at[idx, "借方金額"] = new_val

    # 3. 科目代碼修正
    def fix_account_code(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        # O -> 0
        val = val.replace("O", "0").replace("o", "0")
        return val

    for idx, val in df["科目代碼"].items():
        new_val = fix_account_code(val)
        if new_val != str(val).strip():
            cleaning_log.append({
                "欄位": "科目代碼", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "科目代碼格式修正"
            })
            df.at[idx, "科目代碼"] = new_val

    # 4. 部門名稱修正
    def fix_department(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        val = re.sub(r"\s+", "", val)  # 移除所有空白
        if val and not val.endswith("部"):
            val = val + "部"
        return val

    for idx, val in df["部門"].items():
        new_val = fix_department(val)
        if new_val != str(val):
            cleaning_log.append({
                "欄位": "部門", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "部門名稱標準化"
            })
            df.at[idx, "部門"] = new_val

    # 5. 核准狀態標準化
    status_mapping = {
        "approved": "已核准",
        "核准": "已核准",
        "待核准": "待核准",
        "已核准": "已核准"
    }
    for idx, val in df["核准狀態"].items():
        if val in status_mapping and status_mapping[val] != val:
            new_val = status_mapping[val]
            cleaning_log.append({
                "欄位": "核准狀態", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "狀態名稱標準化"
            })
            df.at[idx, "核准狀態"] = new_val

    # 6. 移除重複傳票
    duplicates = df[df.duplicated(subset=["傳票編號"], keep=False)]
    df = df.drop_duplicates(subset=["傳票編號"], keep="first")
    for idx in duplicates.index[1::2]:  # 每對重複的第二筆
        cleaning_log.append({
            "欄位": "傳票編號", "列號": idx + 2,
            "原始值": duplicates.loc[idx, "傳票編號"], "修正值": "(已刪除)", "原因": "重複傳票"
        })

    # 輸出清理後資料
    df.to_excel(OUTPUT_DIR / "cleaned_expense_detail.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaned_expense_detail.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaned_expense_detail.xlsx")

    # 輸出清理日誌
    log_df = pd.DataFrame(cleaning_log)
    log_df.to_excel(OUTPUT_DIR / "cleaning_log.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaning_log.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaning_log.xlsx")

    print(f"  ✅ 清理完成：{len(cleaning_log)} 處修正")
    print(f"  📄 輸出：cleaned_expense_detail.xlsx")
    print(f"  📄 輸出：cleaning_log.xlsx")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# Task 2: 損益表彙整
# ══════════════════════════════════════════════════════════════════════════════
def generate_income_statement(expense_df):
    """產生年度損益表"""
    print("\n📋 Task 2: 損益表彙整")
    print("-" * 40)

    # 讀取科目對照表
    account_df = pd.read_excel(RAW_DIR / "account_chart.xlsx")

    # 合併科目類別
    expense_df = expense_df.merge(
        account_df[["科目代碼", "科目類別"]],
        on="科目代碼",
        how="left"
    )

    # 轉換金額為數值
    expense_df["借方金額"] = pd.to_numeric(expense_df["借方金額"], errors="coerce").fillna(0)

    # 依月份和類別彙總
    expense_df["月份"] = expense_df["日期"].str.extract(r"/(\d+)/")[0].astype(int)

    # 月度彙總
    monthly_summary = expense_df.groupby(["月份", "科目類別"])["借方金額"].sum().unstack(fill_value=0)
    monthly_summary["合計"] = monthly_summary.sum(axis=1)

    # 類別彙總
    category_summary = expense_df.groupby("科目類別")["借方金額"].sum().reset_index()
    category_summary.columns = ["費用類別", "年度金額"]
    category_summary = category_summary.sort_values("年度金額", ascending=False)
    category_summary["佔比"] = (category_summary["年度金額"] / category_summary["年度金額"].sum() * 100).round(1)

    # 建立 Excel 工作簿
    with pd.ExcelWriter(OUTPUT_DIR / "income_statement.xlsx", engine="openpyxl") as writer:
        # Sheet 1: 月度費用明細
        monthly_summary.to_excel(writer, sheet_name="月度費用明細")

        # Sheet 2: 費用類別彙總
        category_summary.to_excel(writer, sheet_name="費用類別彙總", index=False)

        # Sheet 3: 部門費用分析
        dept_summary = expense_df.groupby(["部門", "科目類別"])["借方金額"].sum().unstack(fill_value=0)
        dept_summary["合計"] = dept_summary.sum(axis=1)
        dept_summary.to_excel(writer, sheet_name="部門費用分析")

    # 套用格式
    wb = load_workbook(OUTPUT_DIR / "income_statement.xlsx")
    for ws in wb.worksheets:
        style_header(ws)
        auto_column_width(ws)
        # 金額格式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > 100:
                    cell.number_format = "#,##0"
    wb.save(OUTPUT_DIR / "income_statement.xlsx")

    print(f"  ✅ 損益表產生完成")
    print(f"  📄 輸出：income_statement.xlsx (3 個工作表)")

    return monthly_summary, category_summary


# ══════════════════════════════════════════════════════════════════════════════
# Task 3: 預算達成分析
# ══════════════════════════════════════════════════════════════════════════════
def analyze_budget_achievement(expense_df):
    """分析預算達成率"""
    print("\n📋 Task 3: 預算達成分析")
    print("-" * 40)

    # 讀取預算資料
    budget_df = pd.read_excel(RAW_DIR / "annual_budget_2025.xlsx")
    account_df = pd.read_excel(RAW_DIR / "account_chart.xlsx")

    # 合併科目類別到費用資料
    expense_df = expense_df.merge(
        account_df[["科目代碼", "科目類別"]],
        on="科目代碼",
        how="left"
    )
    expense_df["借方金額"] = pd.to_numeric(expense_df["借方金額"], errors="coerce").fillna(0)

    # 部門×類別實際支出
    actual = expense_df.groupby(["部門", "科目類別"])["借方金額"].sum().reset_index()
    actual.columns = ["部門", "費用類別", "實際支出"]

    # 合併預算與實際
    analysis = budget_df.merge(actual, on=["部門", "費用類別"], how="left")
    analysis["實際支出"] = analysis["實際支出"].fillna(0)
    analysis["達成率"] = (analysis["實際支出"] / analysis["年度預算"] * 100).round(1)
    analysis["差異"] = analysis["年度預算"] - analysis["實際支出"]
    analysis["狀態"] = analysis["達成率"].apply(
        lambda x: "超支" if x > 100 else ("警示" if x > 90 else "正常")
    )

    # 輸出
    analysis.to_excel(OUTPUT_DIR / "budget_analysis.xlsx", index=False)

    # 套用格式和條件格式
    wb = load_workbook(OUTPUT_DIR / "budget_analysis.xlsx")
    ws = wb.active
    style_header(ws)
    auto_column_width(ws)

    # 條件格式：狀態欄
    status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "狀態":
            status_col = idx
            break

    if status_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            if cell.value == "超支":
                cell.fill = ALERT_FILL
            elif cell.value == "警示":
                cell.fill = WARNING_FILL
            else:
                cell.fill = SUCCESS_FILL

    # 金額格式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and abs(cell.value) > 100:
                cell.number_format = "#,##0"

    wb.save(OUTPUT_DIR / "budget_analysis.xlsx")

    # 統計
    over_budget = len(analysis[analysis["狀態"] == "超支"])
    warning = len(analysis[analysis["狀態"] == "警示"])

    print(f"  ✅ 預算分析完成")
    print(f"     超支項目：{over_budget} 項")
    print(f"     警示項目：{warning} 項")
    print(f"  📄 輸出：budget_analysis.xlsx")

    return analysis


# ══════════════════════════════════════════════════════════════════════════════
# Task 4: 財務儀表板
# ══════════════════════════════════════════════════════════════════════════════
def create_dashboard(monthly_summary, category_summary, budget_analysis):
    """建立財務儀表板"""
    print("\n📋 Task 4: 財務儀表板")
    print("-" * 40)

    with pd.ExcelWriter(OUTPUT_DIR / "finance_dashboard.xlsx", engine="openpyxl") as writer:
        # Sheet 1: 月度趨勢
        monthly_summary.to_excel(writer, sheet_name="月度趨勢")

        # Sheet 2: 費用結構
        category_summary.to_excel(writer, sheet_name="費用結構", index=False)

        # Sheet 3: 預算達成
        budget_summary = budget_analysis.groupby("部門").agg({
            "年度預算": "sum",
            "實際支出": "sum"
        }).reset_index()
        budget_summary["達成率"] = (budget_summary["實際支出"] / budget_summary["年度預算"] * 100).round(1)
        budget_summary.to_excel(writer, sheet_name="部門預算達成", index=False)

    # 套用格式與圖表
    wb = load_workbook(OUTPUT_DIR / "finance_dashboard.xlsx")

    for ws in wb.worksheets:
        style_header(ws)
        auto_column_width(ws)

    # 月度趨勢圖表
    ws = wb["月度趨勢"]
    chart = LineChart()
    chart.title = "月度費用趨勢"
    chart.style = 10
    chart.x_axis.title = "月份"
    chart.y_axis.title = "金額"

    # 找到「合計」欄
    total_col = ws.max_column
    data = Reference(ws, min_col=total_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "A15")

    # 部門預算達成圖表
    ws = wb["部門預算達成"]
    chart = BarChart()
    chart.title = "部門預算 vs 實際支出"
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 10

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "F2")

    wb.save(OUTPUT_DIR / "finance_dashboard.xlsx")

    print(f"  ✅ 儀表板建立完成")
    print(f"  📄 輸出：finance_dashboard.xlsx (含圖表)")


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("📊 財務報表自動化 — 主處理腳本")
    print("=" * 60)

    # Task 1: 資料清理
    expense_df = clean_expense_data()

    # Task 2: 損益表彙整
    monthly_summary, category_summary = generate_income_statement(expense_df)

    # Task 3: 預算達成分析
    budget_analysis = analyze_budget_achievement(expense_df)

    # Task 4: 財務儀表板
    create_dashboard(monthly_summary, category_summary, budget_analysis)

    print("\n" + "=" * 60)
    print("✅ 所有財務報表處理完成！")
    print(f"📁 輸出目錄: {OUTPUT_DIR}")
    print("=" * 60)
