#!/usr/bin/env python3
"""
08-inventory: 庫存管理與自動補貨提醒 — 主處理腳本
讀取庫存資料，產生補貨建議與庫存報表
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference
import re

# ══════════════════════════════════════════════════════════════════════════════
# 路徑設定
# ══════════════════════════════════════════════════════════════════════════════
RAW_DIR = Path(__file__).parent / "raw"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# 樣式定義
# ══════════════════════════════════════════════════════════════════════════════
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def style_header(ws, fill=HEADER_FILL):
    """套用標題列樣式"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """自動調整欄寬"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


# ══════════════════════════════════════════════════════════════════════════════
# Task 1: 資料清理
# ══════════════════════════════════════════════════════════════════════════════
def clean_inventory_data():
    """清理庫存資料"""
    print("\n📋 Task 1: 資料清理")
    print("-" * 40)

    inv_df = pd.read_excel(RAW_DIR / "inventory_detail.xlsx", dtype=str)
    cleaning_log = []

    # 1. 庫存數量轉數值並修正負數
    inv_df["庫存數量"] = pd.to_numeric(inv_df["庫存數量"], errors="coerce").fillna(0).astype(int)
    for idx, val in enumerate(inv_df["庫存數量"]):
        if val < 0:
            cleaning_log.append({
                "欄位": "庫存數量", "列號": idx + 2,
                "原始值": val, "修正值": 0, "原因": "負數庫存修正為0"
            })
            inv_df.at[idx, "庫存數量"] = 0

    # 2. 已預留數量不能大於庫存數量
    inv_df["已預留數量"] = pd.to_numeric(inv_df["已預留數量"], errors="coerce").fillna(0).astype(int)
    for idx, row in inv_df.iterrows():
        if row["已預留數量"] > row["庫存數量"]:
            cleaning_log.append({
                "欄位": "已預留數量", "列號": idx + 2,
                "原始值": row["已預留數量"], "修正值": row["庫存數量"],
                "原因": "預留數量超過庫存，修正為庫存數量"
            })
            inv_df.at[idx, "已預留數量"] = row["庫存數量"]

    # 3. 商品編號標準化
    def fix_sku(val):
        if pd.isna(val):
            return val
        val = str(val).strip().upper()
        val = val.replace(" ", "")
        return val

    for idx, val in inv_df["商品編號"].items():
        new_val = fix_sku(val)
        if new_val != val:
            cleaning_log.append({
                "欄位": "商品編號", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "商品編號格式標準化"
            })
            inv_df.at[idx, "商品編號"] = new_val

    # 4. 日期格式標準化
    def fix_date(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        patterns = [
            (r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", r"\1/\2/\3"),
            (r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})", r"\3/\2/\1"),
        ]
        for pattern, replacement in patterns:
            if re.match(pattern, val):
                val = re.sub(pattern, replacement, val)
                break
        return val

    for idx, val in inv_df["最後盤點日期"].items():
        new_val = fix_date(val)
        if new_val != val:
            cleaning_log.append({
                "欄位": "最後盤點日期", "列號": idx + 2,
                "原始值": val, "修正值": new_val, "原因": "日期格式標準化"
            })
            inv_df.at[idx, "最後盤點日期"] = new_val

    # 輸出清理後資料
    inv_df.to_excel(OUTPUT_DIR / "cleaned_inventory.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaned_inventory.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaned_inventory.xlsx")

    # 輸出清理日誌
    log_df = pd.DataFrame(cleaning_log)
    log_df.to_excel(OUTPUT_DIR / "cleaning_log.xlsx", index=False)
    wb = load_workbook(OUTPUT_DIR / "cleaning_log.xlsx")
    style_header(wb.active)
    auto_column_width(wb.active)
    wb.save(OUTPUT_DIR / "cleaning_log.xlsx")

    print(f"  ✅ 清理完成：{len(cleaning_log)} 處修正")
    print(f"  📄 輸出：cleaned_inventory.xlsx")
    print(f"  📄 輸出：cleaning_log.xlsx")

    return inv_df


# ══════════════════════════════════════════════════════════════════════════════
# Task 2: 庫存水位分析與補貨建議
# ══════════════════════════════════════════════════════════════════════════════
def analyze_and_suggest_reorder(inv_df):
    """分析庫存水位並產生補貨建議"""
    print("\n📋 Task 2: 補貨建議")
    print("-" * 40)

    # 讀取商品主檔
    product_df = pd.read_excel(RAW_DIR / "product_master.xlsx")

    # 彙總各商品總庫存
    inv_summary = inv_df.groupby("商品編號").agg({
        "庫存數量": "sum",
        "已預留數量": "sum"
    }).reset_index()
    inv_summary.columns = ["商品編號", "總庫存", "總預留"]
    inv_summary["可用庫存"] = inv_summary["總庫存"] - inv_summary["總預留"]

    # 合併商品主檔
    analysis = product_df.merge(inv_summary, on="商品編號", how="left")
    analysis["總庫存"] = analysis["總庫存"].fillna(0).astype(int)
    analysis["可用庫存"] = analysis["可用庫存"].fillna(0).astype(int)

    # 庫存水位判定
    def assess_level(row):
        if row["可用庫存"] <= 0:
            return "缺貨"
        elif row["可用庫存"] < row["安全庫存量"]:
            return "低於安全庫存"
        elif row["可用庫存"] < row["安全庫存量"] * 1.5:
            return "庫存偏低"
        else:
            return "正常"

    analysis["庫存狀態"] = analysis.apply(assess_level, axis=1)

    # 補貨建議數量
    def calc_reorder_qty(row):
        if row["庫存狀態"] in ["缺貨", "低於安全庫存"]:
            # 建議補到安全庫存的 2 倍
            target = row["安全庫存量"] * 2
            need = target - row["可用庫存"]
            # 向上取到最低訂購量的倍數
            min_order = row["最低訂購量"]
            return int(np.ceil(need / min_order) * min_order)
        return 0

    analysis["建議補貨數量"] = analysis.apply(calc_reorder_qty, axis=1)
    analysis["預估金額"] = analysis["建議補貨數量"] * analysis["單價"]

    # 輸出補貨建議
    reorder_df = analysis[analysis["建議補貨數量"] > 0][[
        "商品編號", "商品名稱", "類別", "供應商名稱",
        "可用庫存", "安全庫存量", "庫存狀態", "建議補貨數量", "預估金額"
    ]].copy()

    reorder_df.to_excel(OUTPUT_DIR / "reorder_suggestions.xlsx", index=False)

    wb = load_workbook(OUTPUT_DIR / "reorder_suggestions.xlsx")
    ws = wb.active
    style_header(ws)
    auto_column_width(ws)

    # 條件格式：庫存狀態
    status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "庫存狀態":
            status_col = idx
            break

    if status_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            if cell.value == "缺貨":
                cell.fill = ALERT_FILL
            elif cell.value == "低於安全庫存":
                cell.fill = WARNING_FILL

    # 金額格式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 100:
                cell.number_format = "#,##0"

    wb.save(OUTPUT_DIR / "reorder_suggestions.xlsx")

    # 統計
    out_of_stock = len(analysis[analysis["庫存狀態"] == "缺貨"])
    low_stock = len(analysis[analysis["庫存狀態"] == "低於安全庫存"])

    print(f"  缺貨商品：{out_of_stock} 項")
    print(f"  低於安全庫存：{low_stock} 項")
    print(f"  📄 輸出：reorder_suggestions.xlsx ({len(reorder_df)} 項需補貨)")

    return analysis


# ══════════════════════════════════════════════════════════════════════════════
# Task 3: 庫存報表
# ══════════════════════════════════════════════════════════════════════════════
def generate_inventory_report(analysis, inv_df):
    """產生庫存報表"""
    print("\n📋 Task 3: 庫存報表")
    print("-" * 40)

    with pd.ExcelWriter(OUTPUT_DIR / "inventory_report.xlsx", engine="openpyxl") as writer:
        # Sheet 1: 庫存總覽
        summary = analysis[[
            "商品編號", "商品名稱", "類別", "總庫存", "可用庫存",
            "安全庫存量", "庫存狀態"
        ]].copy()
        summary.to_excel(writer, sheet_name="庫存總覽", index=False)

        # Sheet 2: 類別統計
        category_stats = analysis.groupby("類別").agg({
            "商品編號": "count",
            "總庫存": "sum",
            "可用庫存": "sum",
            "預估金額": "sum"
        }).reset_index()
        category_stats.columns = ["類別", "商品數", "總庫存", "可用庫存", "庫存價值"]
        # 計算庫存價值（庫存數量 × 單價）
        value_calc = analysis.groupby("類別").apply(
            lambda x: (x["總庫存"] * x["單價"]).sum()
        ).reset_index()
        value_calc.columns = ["類別", "庫存價值"]
        category_stats = category_stats.drop(columns=["庫存價值"]).merge(value_calc, on="類別")
        category_stats.to_excel(writer, sheet_name="類別統計", index=False)

        # Sheet 3: 倉庫分佈
        warehouse_stats = inv_df.groupby("倉庫").agg({
            "商品編號": "nunique",
            "庫存數量": "sum"
        }).reset_index()
        warehouse_stats.columns = ["倉庫", "商品種類數", "總庫存數量"]
        warehouse_stats.to_excel(writer, sheet_name="倉庫分佈", index=False)

        # Sheet 4: 庫存狀態分佈
        status_stats = analysis["庫存狀態"].value_counts().reset_index()
        status_stats.columns = ["庫存狀態", "商品數"]
        status_stats.to_excel(writer, sheet_name="狀態分佈", index=False)

    # 套用格式
    wb = load_workbook(OUTPUT_DIR / "inventory_report.xlsx")
    for ws in wb.worksheets:
        style_header(ws)
        auto_column_width(ws)
        # 數字格式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > 100:
                    cell.number_format = "#,##0"

    # 狀態顏色
    ws = wb["庫存總覽"]
    status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "庫存狀態":
            status_col = idx
            break

    if status_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            if cell.value == "缺貨":
                cell.fill = ALERT_FILL
            elif cell.value == "低於安全庫存":
                cell.fill = WARNING_FILL
            elif cell.value == "正常":
                cell.fill = SUCCESS_FILL

    wb.save(OUTPUT_DIR / "inventory_report.xlsx")

    print(f"  📄 輸出：inventory_report.xlsx (4 個工作表)")


# ══════════════════════════════════════════════════════════════════════════════
# Task 4: 供應商採購單
# ══════════════════════════════════════════════════════════════════════════════
def generate_purchase_orders(analysis):
    """依供應商產生採購單"""
    print("\n📋 Task 4: 供應商採購單")
    print("-" * 40)

    # 篩選需要補貨的商品
    reorder = analysis[analysis["建議補貨數量"] > 0].copy()

    if len(reorder) == 0:
        print("  ✅ 目前無需補貨")
        return

    # 讀取供應商資料
    supplier_df = pd.read_excel(RAW_DIR / "suppliers.xlsx")

    # 依供應商分組
    for supplier_name in reorder["供應商名稱"].unique():
        supplier_orders = reorder[reorder["供應商名稱"] == supplier_name][[
            "商品編號", "商品名稱", "建議補貨數量", "單價", "預估金額"
        ]].copy()

        # 取得供應商資料
        supplier_info = supplier_df[supplier_df["供應商名稱"] == supplier_name].iloc[0]

        # 建立採購單
        po_file = OUTPUT_DIR / f"PO_{supplier_info['供應商代碼']}.xlsx"
        supplier_orders.to_excel(po_file, index=False, startrow=4)

        wb = load_workbook(po_file)
        ws = wb.active

        # 加入抬頭資訊
        ws["A1"] = "採購單"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"供應商：{supplier_name}"
        ws["A3"] = f"聯絡人：{supplier_info['聯絡人']}  電話：{supplier_info['聯絡電話']}"
        ws["A4"] = f"日期：{datetime.now().strftime('%Y/%m/%d')}"

        style_header(ws, fill=HEADER_FILL)
        # 標題列在第5列
        for cell in ws[5]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        auto_column_width(ws)

        # 加入總計
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=4, value="總計")
        ws.cell(row=total_row, column=5, value=supplier_orders["預估金額"].sum())
        ws.cell(row=total_row, column=5).number_format = "#,##0"
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        ws.cell(row=total_row, column=5).font = Font(bold=True)

        wb.save(po_file)
        print(f"  📄 輸出：PO_{supplier_info['供應商代碼']}.xlsx ({len(supplier_orders)} 項)")


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("📊 庫存管理與自動補貨提醒 — 主處理腳本")
    print("=" * 60)

    # Task 1: 資料清理
    inv_df = clean_inventory_data()

    # Task 2: 庫存分析與補貨建議
    analysis = analyze_and_suggest_reorder(inv_df)

    # Task 3: 庫存報表
    generate_inventory_report(analysis, inv_df)

    # Task 4: 供應商採購單
    generate_purchase_orders(analysis)

    print("\n" + "=" * 60)
    print("✅ 庫存管理處理完成！")
    print(f"📁 輸出目錄: {OUTPUT_DIR}")
    print("=" * 60)
